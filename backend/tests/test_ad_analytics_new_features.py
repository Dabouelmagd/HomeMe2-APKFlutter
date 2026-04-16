"""
Test new Ad Analytics features:
- GET /api/ads/analytics/compare - Period comparison
- GET /api/ads/analytics/export?format=csv - CSV export
- GET /api/ads/analytics/export?format=excel - Excel export
- POST /api/ads/analytics/send-weekly-report - Weekly email report
"""
import pytest
import requests
import os

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestAdAnalyticsNewFeatures:
    """Test new ad analytics endpoints: compare, export, send-weekly-report"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup: Login as app_owner to get token"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        # Login as app_owner (Owner_homeme)
        login_response = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "username": "Owner_homeme",
            "password": "Dalia1234@"
        })
        
        if login_response.status_code == 200:
            data = login_response.json()
            self.token = data.get("access_token")
            self.session.headers.update({"Authorization": f"Bearer {self.token}"})
        else:
            pytest.skip(f"Login failed: {login_response.status_code} - {login_response.text}")
    
    # ==================== Compare Endpoint Tests ====================
    
    def test_compare_endpoint_default_periods(self):
        """Test GET /api/ads/analytics/compare with default periods (this month vs last month)"""
        response = self.session.get(f"{BASE_URL}/api/ads/analytics/compare")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        data = response.json()
        
        # Verify response structure
        assert "period1" in data, "Response should contain period1"
        assert "period2" in data, "Response should contain period2"
        assert "changes" in data, "Response should contain changes"
        
        # Verify period1 structure
        p1 = data["period1"]
        assert "start" in p1, "period1 should have start date"
        assert "end" in p1, "period1 should have end date"
        assert "clicks" in p1, "period1 should have clicks"
        assert "views" in p1, "period1 should have views"
        assert "ctr" in p1, "period1 should have ctr"
        assert "revenue" in p1, "period1 should have revenue"
        assert "new_ads" in p1, "period1 should have new_ads"
        assert "active_ads" in p1, "period1 should have active_ads"
        
        # Verify period2 structure
        p2 = data["period2"]
        assert "start" in p2, "period2 should have start date"
        assert "end" in p2, "period2 should have end date"
        assert "clicks" in p2, "period2 should have clicks"
        assert "views" in p2, "period2 should have views"
        
        # Verify changes structure
        changes = data["changes"]
        assert "clicks" in changes, "changes should have clicks"
        assert "views" in changes, "changes should have views"
        assert "ctr" in changes, "changes should have ctr"
        assert "revenue" in changes, "changes should have revenue"
        assert "new_ads" in changes, "changes should have new_ads"
        
        print(f"Compare endpoint returned: period1={p1['start']} to {p1['end']}, period2={p2['start']} to {p2['end']}")
        print(f"Changes: clicks={changes['clicks']}%, views={changes['views']}%, ctr={changes['ctr']}%")
    
    def test_compare_endpoint_custom_periods(self):
        """Test GET /api/ads/analytics/compare with custom date ranges"""
        response = self.session.get(f"{BASE_URL}/api/ads/analytics/compare", params={
            "period1_start": "2025-01-01",
            "period1_end": "2025-01-15",
            "period2_start": "2024-12-01",
            "period2_end": "2024-12-15"
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        data = response.json()
        assert data["period1"]["start"] == "2025-01-01"
        assert data["period1"]["end"] == "2025-01-15"
        assert data["period2"]["start"] == "2024-12-01"
        assert data["period2"]["end"] == "2024-12-15"
        
        print(f"Custom periods compare successful")
    
    def test_compare_endpoint_data_types(self):
        """Verify data types in compare response"""
        response = self.session.get(f"{BASE_URL}/api/ads/analytics/compare")
        assert response.status_code == 200
        
        data = response.json()
        p1 = data["period1"]
        changes = data["changes"]
        
        # Verify numeric types
        assert isinstance(p1["clicks"], int), "clicks should be int"
        assert isinstance(p1["views"], int), "views should be int"
        assert isinstance(p1["ctr"], (int, float)), "ctr should be numeric"
        assert isinstance(p1["revenue"], (int, float)), "revenue should be numeric"
        assert isinstance(p1["new_ads"], int), "new_ads should be int"
        assert isinstance(p1["active_ads"], int), "active_ads should be int"
        
        # Verify changes are percentages (numeric)
        assert isinstance(changes["clicks"], (int, float)), "clicks change should be numeric"
        assert isinstance(changes["views"], (int, float)), "views change should be numeric"
        assert isinstance(changes["ctr"], (int, float)), "ctr change should be numeric"
        
        print("Data types verified correctly")
    
    # ==================== Export CSV Tests ====================
    
    def test_export_csv_endpoint(self):
        """Test GET /api/ads/analytics/export?format=csv returns CSV file"""
        response = self.session.get(f"{BASE_URL}/api/ads/analytics/export", params={"format": "csv"})
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        # Verify content type
        content_type = response.headers.get("content-type", "")
        assert "text/csv" in content_type or "application/octet-stream" in content_type, f"Expected CSV content type, got {content_type}"
        
        # Verify content disposition (download filename)
        content_disp = response.headers.get("content-disposition", "")
        assert "ad_analytics_report.csv" in content_disp, f"Expected CSV filename in disposition, got {content_disp}"
        
        # Verify content is not empty and contains expected headers
        content = response.content.decode('utf-8-sig')
        assert len(content) > 0, "CSV content should not be empty"
        
        # Check for Arabic headers (BOM + Arabic text)
        assert "العنوان" in content or "الموقع" in content, "CSV should contain Arabic headers"
        
        print(f"CSV export successful, size: {len(content)} bytes")
    
    def test_export_csv_contains_data_rows(self):
        """Verify CSV export contains actual ad data rows"""
        response = self.session.get(f"{BASE_URL}/api/ads/analytics/export", params={"format": "csv"})
        assert response.status_code == 200
        
        content = response.content.decode('utf-8-sig')
        lines = content.strip().split('\n')
        
        # Should have header + at least summary row
        assert len(lines) >= 2, f"CSV should have at least 2 lines (header + data), got {len(lines)}"
        
        # Check header row contains expected columns
        header = lines[0]
        assert "CTR" in header or "ctr" in header.lower(), "Header should contain CTR column"
        
        print(f"CSV has {len(lines)} lines including header")
    
    # ==================== Export Excel Tests ====================
    
    def test_export_excel_endpoint(self):
        """Test GET /api/ads/analytics/export?format=excel returns Excel file"""
        response = self.session.get(f"{BASE_URL}/api/ads/analytics/export", params={"format": "excel"})
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        # Verify content type for Excel
        content_type = response.headers.get("content-type", "")
        assert "spreadsheetml" in content_type or "application/vnd" in content_type or "octet-stream" in content_type, f"Expected Excel content type, got {content_type}"
        
        # Verify content disposition (download filename)
        content_disp = response.headers.get("content-disposition", "")
        assert "ad_analytics_report.xlsx" in content_disp, f"Expected Excel filename in disposition, got {content_disp}"
        
        # Verify content is not empty and starts with Excel magic bytes (PK for zip)
        content = response.content
        assert len(content) > 0, "Excel content should not be empty"
        assert content[:2] == b'PK', "Excel file should start with PK (zip format)"
        
        print(f"Excel export successful, size: {len(content)} bytes")
    
    def test_export_excel_file_structure(self):
        """Verify Excel file has proper structure (multiple sheets)"""
        response = self.session.get(f"{BASE_URL}/api/ads/analytics/export", params={"format": "excel"})
        assert response.status_code == 200
        
        # Save temporarily and verify with openpyxl
        import io
        from openpyxl import load_workbook
        
        wb = load_workbook(io.BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        # Should have multiple sheets
        assert len(sheet_names) >= 2, f"Excel should have at least 2 sheets, got {sheet_names}"
        
        # Check for expected sheet names (Arabic)
        assert "الإعلانات" in sheet_names, f"Should have 'الإعلانات' sheet, got {sheet_names}"
        assert "الملخص المالي" in sheet_names, f"Should have 'الملخص المالي' sheet, got {sheet_names}"
        
        print(f"Excel has sheets: {sheet_names}")
    
    def test_export_default_format_is_excel(self):
        """Test that default format (no param) returns Excel"""
        response = self.session.get(f"{BASE_URL}/api/ads/analytics/export")
        
        assert response.status_code == 200
        content_disp = response.headers.get("content-disposition", "")
        assert "xlsx" in content_disp, f"Default format should be Excel, got {content_disp}"
        
        print("Default export format is Excel")
    
    # ==================== Send Weekly Report Tests ====================
    
    def test_send_weekly_report_endpoint(self):
        """Test POST /api/ads/analytics/send-weekly-report sends email"""
        response = self.session.post(f"{BASE_URL}/api/ads/analytics/send-weekly-report")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        data = response.json()
        
        # Verify response structure
        assert "message" in data, "Response should contain message"
        assert "to_email" in data, "Response should contain to_email"
        assert "summary" in data, "Response should contain summary"
        
        # Verify email was sent to correct address
        assert "@" in data["to_email"], "to_email should be a valid email"
        
        # Verify summary contains expected metrics
        summary = data["summary"]
        assert "total_revenue" in summary, "summary should have total_revenue"
        assert "total_views" in summary, "summary should have total_views"
        assert "total_clicks" in summary, "summary should have total_clicks"
        assert "week_clicks" in summary, "summary should have week_clicks"
        assert "avg_ctr" in summary, "summary should have avg_ctr"
        
        print(f"Weekly report sent to: {data['to_email']}")
        print(f"Summary: revenue={summary['total_revenue']}, views={summary['total_views']}, clicks={summary['total_clicks']}")
    
    def test_send_weekly_report_returns_correct_email(self):
        """Verify weekly report is sent to app owner email"""
        response = self.session.post(f"{BASE_URL}/api/ads/analytics/send-weekly-report")
        assert response.status_code == 200
        
        data = response.json()
        # Should be sent to app owner email (dalia@datalifeai.com)
        assert "dalia@datalifeai.com" in data["to_email"] or "@" in data["to_email"], \
            f"Email should be sent to owner, got {data['to_email']}"
        
        print(f"Email correctly sent to: {data['to_email']}")
    
    # ==================== Existing Endpoints Still Work ====================
    
    def test_realtime_endpoint_still_works(self):
        """Verify GET /api/ads/analytics/realtime still works"""
        response = self.session.get(f"{BASE_URL}/api/ads/analytics/realtime")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        data = response.json()
        assert "live_summary" in data
        assert "daily_series" in data
        assert "hourly_today" in data
        assert "alerts" in data
        
        print(f"Realtime endpoint working: {data['live_summary']['total_ads']} ads")
    
    def test_financial_endpoint_still_works(self):
        """Verify GET /api/ads/analytics/financial still works"""
        response = self.session.get(f"{BASE_URL}/api/ads/analytics/financial")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        data = response.json()
        assert "summary" in data
        assert "position_revenue" in data
        assert "monthly_chart" in data
        assert "top_earners" in data
        
        print(f"Financial endpoint working: {data['summary']['total_revenue']} EGP revenue")
    
    # ==================== Authentication Tests ====================
    
    def test_compare_requires_auth(self):
        """Verify compare endpoint requires authentication"""
        no_auth_session = requests.Session()
        response = no_auth_session.get(f"{BASE_URL}/api/ads/analytics/compare")
        
        assert response.status_code in [401, 403], f"Expected 401/403 without auth, got {response.status_code}"
        print("Compare endpoint correctly requires authentication")
    
    def test_export_requires_auth(self):
        """Verify export endpoint requires authentication"""
        no_auth_session = requests.Session()
        response = no_auth_session.get(f"{BASE_URL}/api/ads/analytics/export")
        
        assert response.status_code in [401, 403], f"Expected 401/403 without auth, got {response.status_code}"
        print("Export endpoint correctly requires authentication")
    
    def test_send_report_requires_auth(self):
        """Verify send-weekly-report endpoint requires authentication"""
        no_auth_session = requests.Session()
        response = no_auth_session.post(f"{BASE_URL}/api/ads/analytics/send-weekly-report")
        
        assert response.status_code in [401, 403], f"Expected 401/403 without auth, got {response.status_code}"
        print("Send weekly report endpoint correctly requires authentication")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])

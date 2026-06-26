"""
Internal Ads Management System - Super Admin controlled
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from datetime import datetime, timezone, timedelta
from typing import Optional
from pydantic import BaseModel
import uuid
import logging
import os
import asyncio
import base64
import httpx

from database import get_db
from auth_deps import get_current_user, require_super_admin

router = APIRouter(prefix="/api")


class AdCreate(BaseModel):
    title: str
    image_url: str = ""
    media_type: str = "image"  # image | video
    template_style: Optional[str] = None  # e.g. 'purple_dream' (used when no image)
    link_url: str = ""
    description: str = ""
    position: str = "sidebar"  # sidebar, banner, inline, dashboard, homepage_hero, homepage_mid, homepage_footer, login_page, popup, notification, splash, services_page
    dimensions: str = ""  # e.g. "728x90", "300x250"
    target_compounds: list = []  # empty = all compounds
    is_active: bool = True
    is_gift: bool = False
    ad_value: float = 0  # monetary value (0 if gift)
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    priority: int = 0


class AdUpdate(BaseModel):
    title: Optional[str] = None
    image_url: Optional[str] = None
    media_type: Optional[str] = None
    template_style: Optional[str] = None
    link_url: Optional[str] = None
    description: Optional[str] = None
    position: Optional[str] = None
    dimensions: Optional[str] = None
    target_compounds: Optional[list] = None
    is_active: Optional[bool] = None
    is_gift: Optional[bool] = None
    ad_value: Optional[float] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    priority: Optional[int] = None
    # 🛡️ Opt-in flag required to actually clear image_url. Prevents accidental
    # wipe when the UI sends image_url="" without user intent.
    clear_image: Optional[bool] = None


# ==================== Ad Campaigns ====================

class CampaignCreate(BaseModel):
    name: str
    description: str = ""
    start_date: str
    end_date: str
    budget: float = 0
    status: str = "draft"  # draft, active, paused, completed, cancelled
    ad_ids: list = []
    positions: list = []
    auto_renew: bool = False
    free_trial_days: int = 0
    target_compounds: list = []

class CampaignUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    budget: Optional[float] = None
    status: Optional[str] = None
    ad_ids: Optional[list] = None
    positions: Optional[list] = None
    auto_renew: Optional[bool] = None
    free_trial_days: Optional[int] = None
    target_compounds: Optional[list] = None


@router.get("/ads/campaigns")
async def get_campaigns(current_user: dict = Depends(require_super_admin)):
    db = get_db()
    campaigns = await db.ad_campaigns.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    # Enrich with ad details
    for c in campaigns:
        ad_ids = c.get("ad_ids", [])
        if ad_ids:
            c_ads = await db.internal_ads.find({"id": {"$in": ad_ids}}, {"_id": 0, "id": 1, "title": 1, "position": 1, "is_active": 1, "views": 1, "clicks": 1}).to_list(50)
            c["ads"] = c_ads
            c["total_views"] = sum(a.get("views", 0) for a in c_ads)
            c["total_clicks"] = sum(a.get("clicks", 0) for a in c_ads)
        else:
            c["ads"] = []
            c["total_views"] = 0
            c["total_clicks"] = 0
    
    stats = {
        "total": len(campaigns),
        "active": len([c for c in campaigns if c.get("status") == "active"]),
        "draft": len([c for c in campaigns if c.get("status") == "draft"]),
        "completed": len([c for c in campaigns if c.get("status") == "completed"]),
        "total_budget": sum(c.get("budget", 0) for c in campaigns),
    }
    return {"campaigns": campaigns, "stats": stats}


@router.post("/ads/campaigns")
async def create_campaign(data: CampaignCreate, current_user: dict = Depends(require_super_admin)):
    db = get_db()
    campaign = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "description": data.description,
        "start_date": data.start_date,
        "end_date": data.end_date,
        "budget": data.budget,
        "status": data.status,
        "ad_ids": data.ad_ids,
        "positions": data.positions,
        "auto_renew": data.auto_renew,
        "free_trial_days": data.free_trial_days,
        "target_compounds": data.target_compounds,
        "created_by": current_user.get("username", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.ad_campaigns.insert_one(campaign)
    del campaign["_id"]
    
    # Activate ads in this campaign if status is active
    if data.status == "active" and data.ad_ids:
        await db.internal_ads.update_many(
            {"id": {"$in": data.ad_ids}},
            {"$set": {"is_active": True, "start_date": data.start_date, "end_date": data.end_date}}
        )
    
    return {"message": "تم إنشاء الحملة الإعلانية", "campaign": campaign}


@router.put("/ads/campaigns/{campaign_id}")
async def update_campaign(campaign_id: str, data: CampaignUpdate, current_user: dict = Depends(require_super_admin)):
    db = get_db()
    update = {"updated_at": datetime.now(timezone.utc).isoformat()}
    for field in ["name", "description", "start_date", "end_date", "budget", "status", "ad_ids", "positions", "auto_renew", "free_trial_days", "target_compounds"]:
        val = getattr(data, field, None)
        if val is not None:
            update[field] = val
    
    await db.ad_campaigns.update_one({"id": campaign_id}, {"$set": update})
    
    # Handle status changes
    if data.status == "active" and data.ad_ids:
        await db.internal_ads.update_many({"id": {"$in": data.ad_ids}}, {"$set": {"is_active": True}})
    elif data.status == "paused" and data.ad_ids:
        await db.internal_ads.update_many({"id": {"$in": data.ad_ids}}, {"$set": {"is_active": False}})
    elif data.status == "cancelled":
        campaign = await db.ad_campaigns.find_one({"id": campaign_id})
        if campaign and campaign.get("ad_ids"):
            await db.internal_ads.update_many({"id": {"$in": campaign["ad_ids"]}}, {"$set": {"is_active": False}})
    
    return {"message": "تم تحديث الحملة"}


@router.post("/ads/campaigns/{campaign_id}/renew")
async def renew_campaign(campaign_id: str, body: dict, current_user: dict = Depends(require_super_admin)):
    db = get_db()
    campaign = await db.ad_campaigns.find_one({"id": campaign_id}, {"_id": 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="الحملة غير موجودة")
    
    new_start = body.get("new_start_date", campaign.get("end_date", ""))
    new_end = body.get("new_end_date", "")
    if not new_end:
        # Default: same duration
        from datetime import datetime as dt
        try:
            s = dt.fromisoformat(campaign["start_date"])
            e = dt.fromisoformat(campaign["end_date"])
            duration = (e - s).days
            ns = dt.fromisoformat(new_start)
            new_end = (ns + timedelta(days=duration)).strftime("%Y-%m-%d")
        except:
            new_end = new_start
    
    await db.ad_campaigns.update_one({"id": campaign_id}, {"$set": {
        "start_date": new_start,
        "end_date": new_end,
        "status": "active",
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }})
    
    # Reactivate ads
    if campaign.get("ad_ids"):
        await db.internal_ads.update_many(
            {"id": {"$in": campaign["ad_ids"]}},
            {"$set": {"is_active": True, "start_date": new_start, "end_date": new_end}}
        )
    
    return {"message": "تم تجديد الحملة", "new_start": new_start, "new_end": new_end}


@router.delete("/ads/campaigns/{campaign_id}")
async def delete_campaign(campaign_id: str, current_user: dict = Depends(require_super_admin)):
    db = get_db()
    campaign = await db.ad_campaigns.find_one({"id": campaign_id}, {"_id": 0})
    if campaign and campaign.get("ad_ids"):
        await db.internal_ads.update_many({"id": {"$in": campaign["ad_ids"]}}, {"$set": {"is_active": False}})
    await db.ad_campaigns.delete_one({"id": campaign_id})
    return {"message": "تم حذف الحملة"}


@router.post("/ads")
async def create_ad(data: AdCreate, current_user: dict = Depends(require_super_admin)):
    db = get_db()
    ad = {
        "id": str(uuid.uuid4()),
        "title": data.title,
        "image_url": data.image_url,
        "media_type": data.media_type or "image",
        "template_style": data.template_style or "purple_dream",
        "link_url": data.link_url,
        "description": data.description,
        "position": data.position,
        "dimensions": data.dimensions,
        "target_compounds": data.target_compounds,
        "is_active": data.is_active,
        "is_gift": data.is_gift,
        "ad_value": data.ad_value if not data.is_gift else 0,
        "start_date": data.start_date,
        "end_date": data.end_date,
        "priority": data.priority,
        "clicks": 0,
        "views": 0,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.internal_ads.insert_one(ad)
    ad.pop("_id", None)
    return {"message": "تم إنشاء الإعلان بنجاح", "ad": ad}


@router.get("/ads")
async def get_all_ads(current_user: dict = Depends(require_super_admin)):
    db = get_db()
    ads = await db.internal_ads.find({}, {"_id": 0}).sort("priority", -1).to_list(200)
    MAX_SLOTS = {
        "homepage_hero": 3, "homepage_mid": 2, "homepage_footer": 2,
        "banner": 5, "sidebar": 3, "dashboard": 2, "inline": 4,
        "login_page": 2, "popup": 1, "notification": 2, "splash": 1, "services_page": 3,
    }
    booked_by_pos = {}
    for a in ads:
        p = a.get("position", "unknown")
        booked_by_pos[p] = booked_by_pos.get(p, 0) + 1

    slot_stats = {}
    for pos, max_s in MAX_SLOTS.items():
        booked = booked_by_pos.get(pos, 0)
        slot_stats[pos] = {"max_slots": max_s, "booked": booked, "available": max_s - booked}

    stats = {
        "total": len(ads),
        "active": len([a for a in ads if a.get("is_active")]),
        "total_clicks": sum(a.get("clicks", 0) for a in ads),
        "total_views": sum(a.get("views", 0) for a in ads),
        "total_revenue": sum(a.get("ad_value", 0) for a in ads if not a.get("is_gift")),
        "gift_ads": len([a for a in ads if a.get("is_gift")]),
        "total_slots": sum(MAX_SLOTS.values()),
        "total_booked": sum(booked_by_pos.values()),
        "total_available": sum(MAX_SLOTS.values()) - sum(booked_by_pos.values()),
        "slot_stats": slot_stats,
    }
    return {"ads": ads, "stats": stats}


@router.get("/ads/by-compounds")
async def get_ads_by_compounds(compound_ids: str = "", current_user: dict = Depends(require_super_admin)):
    """Get ads that target specific compounds"""
    db = get_db()
    cids = [c.strip() for c in compound_ids.split(",") if c.strip()] if compound_ids else []
    all_ads = await db.internal_ads.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    if cids:
        filtered = [a for a in all_ads if any(c in a.get("target_compounds", []) for c in cids) or (not a.get("target_compounds") and False)]
    else:
        filtered = all_ads
    return {"ads": filtered}


# ==================== Hybrid Ad Settings ====================

@router.get("/ads/ad-settings")
async def get_ad_settings(current_user: dict = Depends(get_current_user)):
    """Get ad display settings for each position"""
    db = get_db()
    settings = await db.app_settings.find_one({"key": "ad_settings"}, {"_id": 0}) or {}
    defaults = {
        "banner": {"mode": "internal_first", "adsense_enabled": True, "internal_enabled": True},
        "sidebar": {"mode": "internal_first", "adsense_enabled": False, "internal_enabled": True},
        "dashboard": {"mode": "internal_only", "adsense_enabled": False, "internal_enabled": True},
        "inline": {"mode": "internal_first", "adsense_enabled": True, "internal_enabled": True},
    }
    positions = settings.get("positions", defaults)
    return {
        "adsense_publisher_id": settings.get("adsense_publisher_id", "ca-pub-5928973437129941"),
        "adsense_global_enabled": settings.get("adsense_global_enabled", True),
        "positions": positions,
    }


@router.put("/ads/ad-settings")
async def update_ad_settings(body: dict, current_user: dict = Depends(require_super_admin)):
    """Update ad display settings"""
    db = get_db()
    update = {"key": "ad_settings"}
    if "adsense_global_enabled" in body:
        update["adsense_global_enabled"] = body["adsense_global_enabled"]
    if "adsense_publisher_id" in body:
        update["adsense_publisher_id"] = body["adsense_publisher_id"]
    if "positions" in body:
        update["positions"] = body["positions"]
    await db.app_settings.update_one({"key": "ad_settings"}, {"$set": update}, upsert=True)
    return {"message": "تم تحديث إعدادات الإعلانات"}


@router.get("/ads/public")
async def get_public_ads(position: str = ""):
    """Get active ads for public pages (no auth required)"""
    db = get_db()
    homepage_positions = ["homepage_hero", "homepage_mid", "homepage_footer", "login_page"]
    query = {"is_active": True}
    if position and position in homepage_positions:
        query["position"] = position
    else:
        query["position"] = {"$in": homepage_positions}

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    ads = await db.internal_ads.find(query, {"_id": 0}).sort("priority", -1).to_list(20)
    valid = []
    for a in ads:
        if a.get("start_date") and a["start_date"] > now:
            continue
        if a.get("end_date") and a["end_date"] < now:
            continue
        valid.append(a)

    # Re-rank: ads with media first (so a single-slot rotator never picks an
    # empty placeholder over a real image). Within each group keep priority/recency order.
    valid.sort(
        key=lambda a: (
            0 if (a.get("image_url") or a.get("video_url") or a.get("media_url")) else 1,
            -int(a.get("priority", 0) or 0),
            -((a.get("created_at") or "")[:19].count("")),
        )
    )

    # Settings
    settings = await db.app_settings.find_one({"key": "ad_settings"}, {"_id": 0}) or {}
    return {"ads": valid, "settings": settings}


@router.get("/ads/active")
async def get_active_ads(position: str = "", compound_id: str = "", current_user: dict = Depends(get_current_user)):
    """Get active ads for display inside the app"""
    db = get_db()
    query = {"is_active": True}
    if position:
        query["position"] = position

    # Check date validity
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    ads = await db.internal_ads.find(query, {"_id": 0}).sort("priority", -1).to_list(50)

    # Filter by compound targeting
    if compound_id:
        ads = [a for a in ads if not a.get("target_compounds") or compound_id in a.get("target_compounds", [])]

    # Filter by date range
    valid_ads = []
    for a in ads:
        start = a.get("start_date")
        end = a.get("end_date")
        if start and start > now:
            continue
        if end and end < now:
            continue
        valid_ads.append(a)

    # Re-rank: ads with media first
    valid_ads.sort(
        key=lambda a: (
            0 if (a.get("image_url") or a.get("video_url") or a.get("media_url")) else 1,
            -int(a.get("priority", 0) or 0),
        )
    )

    # Track views
    ad_ids = [a["id"] for a in valid_ads if a.get("id")]
    if ad_ids:
        await db.internal_ads.update_many(
            {"id": {"$in": ad_ids}},
            {"$inc": {"views": 1}}
        )

    return {"ads": valid_ads}


@router.put("/ads/{ad_id}")
async def update_ad(ad_id: str, data: AdUpdate, current_user: dict = Depends(require_super_admin)):
    db = get_db()
    update = {k: v for k, v in data.dict().items() if v is not None}
    # 🛡️ Protect against accidental image wipe: an empty image_url is treated
    # as "no change" unless the UI sent clear_image=true explicitly.
    clear_image = update.pop("clear_image", False)
    if "image_url" in update and update["image_url"] == "" and not clear_image:
        update.pop("image_url", None)
    if not update:
        raise HTTPException(status_code=400, detail="لا توجد بيانات للتحديث")
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.internal_ads.update_one({"id": ad_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="إعلان غير موجود")
    return {"message": "تم تحديث الإعلان"}


@router.put("/ads/{ad_id}/toggle")
async def toggle_ad(ad_id: str, current_user: dict = Depends(require_super_admin)):
    db = get_db()
    ad = await db.internal_ads.find_one({"id": ad_id})
    if not ad:
        raise HTTPException(status_code=404, detail="إعلان غير موجود")
    new_status = not ad.get("is_active", True)
    await db.internal_ads.update_one({"id": ad_id}, {"$set": {"is_active": new_status}})
    return {"message": f"تم {'تفعيل' if new_status else 'تعطيل'} الإعلان", "is_active": new_status}


@router.delete("/ads/{ad_id}")
async def delete_ad(ad_id: str, current_user: dict = Depends(require_super_admin)):
    db = get_db()
    result = await db.internal_ads.delete_one({"id": ad_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="إعلان غير موجود")
    return {"message": "تم حذف الإعلان"}


@router.post("/ads/{ad_id}/click")
async def track_ad_click(ad_id: str, current_user: dict = Depends(get_current_user)):
    db = get_db()
    now = datetime.now(timezone.utc)
    await db.internal_ads.update_one({"id": ad_id}, {"$inc": {"clicks": 1}})
    # Log click event for analytics
    await db.ad_events.insert_one({
        "ad_id": ad_id,
        "event": "click",
        "user_id": current_user.get("id", ""),
        "compound_id": current_user.get("compound_id", ""),
        "timestamp": now.isoformat(),
    })
    return {"ok": True}


@router.get("/ads/analytics")
async def get_ad_analytics(current_user: dict = Depends(require_super_admin)):
    """Get ad performance analytics for charts"""
    db = get_db()
    ads = await db.internal_ads.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)

    # Per-ad stats
    ad_stats = []
    for a in ads:
        clicks = a.get("clicks", 0)
        views = a.get("views", 0)
        ctr = round((clicks / views * 100), 2) if views > 0 else 0
        ad_stats.append({
            "id": a.get("id", ""),
            "title": a.get("title", ""),
            "position": a.get("position", ""),
            "dimensions": a.get("dimensions", ""),
            "is_gift": a.get("is_gift", False),
            "ad_value": a.get("ad_value", 0),
            "clicks": clicks,
            "views": views,
            "ctr": ctr,
            "is_active": a.get("is_active", False),
            "created_at": a.get("created_at", ""),
        })

    # Sort by CTR descending for top performers
    top_by_ctr = sorted([a for a in ad_stats if a["views"] > 0], key=lambda x: x["ctr"], reverse=True)[:5]
    top_by_clicks = sorted(ad_stats, key=lambda x: x["clicks"], reverse=True)[:5]
    top_by_views = sorted(ad_stats, key=lambda x: x["views"], reverse=True)[:5]

    # Aggregate by position
    position_stats = {}
    for a in ad_stats:
        pos = a["position"]
        if pos not in position_stats:
            position_stats[pos] = {"clicks": 0, "views": 0, "count": 0, "revenue": 0}
        position_stats[pos]["clicks"] += a["clicks"]
        position_stats[pos]["views"] += a["views"]
        position_stats[pos]["count"] += 1
        if not a["is_gift"]:
            position_stats[pos]["revenue"] += a["ad_value"]

    position_chart = [
        {"label": pos, "clicks": s["clicks"], "views": s["views"], "count": s["count"], "revenue": s["revenue"]}
        for pos, s in position_stats.items()
    ]

    # Overall summary
    total_clicks = sum(a["clicks"] for a in ad_stats)
    total_views = sum(a["views"] for a in ad_stats)
    total_revenue = sum(a["ad_value"] for a in ad_stats if not a["is_gift"])
    gift_count = len([a for a in ad_stats if a["is_gift"]])
    avg_ctr = round((total_clicks / total_views * 100), 2) if total_views > 0 else 0

    return {
        "summary": {
            "total_ads": len(ad_stats),
            "active_ads": len([a for a in ad_stats if a["is_active"]]),
            "total_clicks": total_clicks,
            "total_views": total_views,
            "avg_ctr": avg_ctr,
            "total_revenue": total_revenue,
            "gift_ads": gift_count,
        },
        "all_ads": ad_stats,
        "top_by_ctr": top_by_ctr,
        "top_by_clicks": top_by_clicks,
        "top_by_views": top_by_views,
        "position_chart": position_chart,
    }


@router.get("/ads/analytics/realtime")
async def get_realtime_ad_analytics(days: int = 30, current_user: dict = Depends(require_super_admin)):
    """Real-time ad analytics with time-series data and CTR alerts"""
    db = get_db()
    now = datetime.now(timezone.utc)
    start_date = (now - timedelta(days=days)).isoformat()

    # Get all ads
    ads = await db.internal_ads.find({}, {"_id": 0}).to_list(500)

    # Get click events for the period
    events = await db.ad_events.find(
        {"timestamp": {"$gte": start_date}}, {"_id": 0}
    ).sort("timestamp", 1).to_list(10000)

    # Build daily time-series from events
    daily_clicks = {}
    hourly_clicks_today = {}
    today_str = now.strftime("%Y-%m-%d")

    for ev in events:
        ts = ev.get("timestamp", "")
        day = ts[:10] if len(ts) >= 10 else ""
        hour = ts[11:13] if len(ts) >= 13 else ""

        if day:
            daily_clicks.setdefault(day, 0)
            daily_clicks[day] += 1

        if day == today_str and hour:
            hourly_clicks_today.setdefault(hour, 0)
            hourly_clicks_today[hour] += 1

    # Build daily views approximation from ad views / days active
    daily_series = []
    for i in range(min(days, 30)):
        d = (now - timedelta(days=i)).strftime("%Y-%m-%d")
        clicks = daily_clicks.get(d, 0)
        # Estimate daily views based on total views / active days
        total_views_per_day = sum(a.get("views", 0) for a in ads) // max(days, 1)
        daily_series.append({
            "date": d,
            "clicks": clicks,
            "views": total_views_per_day + clicks * 5,
            "ctr": round((clicks / max(total_views_per_day + clicks * 5, 1)) * 100, 2)
        })
    daily_series.reverse()

    # Hourly breakdown for today
    hourly_series = []
    for h in range(24):
        hh = f"{h:02d}"
        hourly_series.append({
            "hour": f"{hh}:00",
            "clicks": hourly_clicks_today.get(hh, 0)
        })

    # CTR alerts - ads with notably high or changing CTR
    alerts = []
    for a in ads:
        views = a.get("views", 0)
        clicks = a.get("clicks", 0)
        if views < 10:
            continue
        ctr = round((clicks / views) * 100, 2)
        if ctr >= 5:
            alerts.append({
                "type": "high_ctr",
                "severity": "success",
                "ad_id": a.get("id"),
                "ad_title": a.get("title"),
                "message": f"CTR عالي: {ctr}%",
                "ctr": ctr,
                "views": views,
                "clicks": clicks
            })
        elif ctr >= 2:
            alerts.append({
                "type": "good_ctr",
                "severity": "info",
                "ad_id": a.get("id"),
                "ad_title": a.get("title"),
                "message": f"CTR جيد: {ctr}%",
                "ctr": ctr,
                "views": views,
                "clicks": clicks
            })

    # Ads with 0 clicks but many views
    for a in ads:
        views = a.get("views", 0)
        clicks = a.get("clicks", 0)
        if views > 50 and clicks == 0 and a.get("is_active"):
            alerts.append({
                "type": "no_clicks",
                "severity": "warning",
                "ad_id": a.get("id"),
                "ad_title": a.get("title"),
                "message": f"لا توجد نقرات رغم {views} مشاهدة",
                "ctr": 0,
                "views": views,
                "clicks": 0
            })

    alerts.sort(key=lambda x: x.get("ctr", 0), reverse=True)

    # Live summary
    total_clicks = sum(a.get("clicks", 0) for a in ads)
    total_views = sum(a.get("views", 0) for a in ads)
    today_clicks = sum(v for k, v in daily_clicks.items() if k == today_str)

    return {
        "live_summary": {
            "total_ads": len(ads),
            "active_ads": len([a for a in ads if a.get("is_active")]),
            "total_views": total_views,
            "total_clicks": total_clicks,
            "today_clicks": today_clicks,
            "avg_ctr": round((total_clicks / max(total_views, 1)) * 100, 2),
            "total_events_period": len(events),
            "last_updated": now.isoformat(),
        },
        "daily_series": daily_series,
        "hourly_today": hourly_series,
        "alerts": alerts[:20],
        "alert_counts": {
            "high_ctr": len([a for a in alerts if a["type"] == "high_ctr"]),
            "good_ctr": len([a for a in alerts if a["type"] == "good_ctr"]),
            "no_clicks": len([a for a in alerts if a["type"] == "no_clicks"]),
        }
    }


@router.get("/ads/analytics/financial")
async def get_financial_ad_analytics(period: str = "monthly", current_user: dict = Depends(require_super_admin)):
    """Detailed financial analytics for ads"""
    db = get_db()
    now = datetime.now(timezone.utc)
    ads = await db.internal_ads.find({}, {"_id": 0}).to_list(500)

    paid_ads = [a for a in ads if not a.get("is_gift")]
    gift_ads = [a for a in ads if a.get("is_gift")]

    # Revenue by position
    rev_by_position = {}
    for a in paid_ads:
        pos = a.get("position", "other")
        rev_by_position.setdefault(pos, {"revenue": 0, "count": 0, "clicks": 0, "views": 0})
        rev_by_position[pos]["revenue"] += a.get("ad_value", 0)
        rev_by_position[pos]["count"] += 1
        rev_by_position[pos]["clicks"] += a.get("clicks", 0)
        rev_by_position[pos]["views"] += a.get("views", 0)

    position_revenue = [
        {"position": k, **v, "cpc": round(v["revenue"] / max(v["clicks"], 1), 2), "cpv": round(v["revenue"] / max(v["views"], 1), 4)}
        for k, v in rev_by_position.items()
    ]

    # Revenue by month (simulate from created_at)
    monthly_revenue = {}
    for a in paid_ads:
        created = a.get("created_at", "")
        month = created[:7] if len(created) >= 7 else "unknown"
        monthly_revenue.setdefault(month, {"revenue": 0, "count": 0})
        monthly_revenue[month]["revenue"] += a.get("ad_value", 0)
        monthly_revenue[month]["count"] += 1

    monthly_chart = sorted([
        {"month": k, "revenue": v["revenue"], "count": v["count"]}
        for k, v in monthly_revenue.items()
    ], key=lambda x: x["month"])

    # Overall financial summary
    total_revenue = sum(a.get("ad_value", 0) for a in paid_ads)
    total_clicks = sum(a.get("clicks", 0) for a in paid_ads)
    total_views = sum(a.get("views", 0) for a in paid_ads)
    avg_ad_value = round(total_revenue / max(len(paid_ads), 1), 2)
    cost_per_click = round(total_revenue / max(total_clicks, 1), 2)
    cost_per_view = round(total_revenue / max(total_views, 1), 4)

    # Current month vs previous month
    current_month = now.strftime("%Y-%m")
    prev_month = (now.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
    current_rev = monthly_revenue.get(current_month, {}).get("revenue", 0)
    prev_rev = monthly_revenue.get(prev_month, {}).get("revenue", 0)
    growth = round(((current_rev - prev_rev) / max(prev_rev, 1)) * 100, 1) if prev_rev else 0

    # Top earning ads
    top_earners = sorted(paid_ads, key=lambda x: x.get("ad_value", 0), reverse=True)[:10]
    top_earners_clean = [{
        "id": a.get("id"), "title": a.get("title"), "position": a.get("position"),
        "ad_value": a.get("ad_value", 0), "clicks": a.get("clicks", 0),
        "views": a.get("views", 0),
        "cpc": round(a.get("ad_value", 0) / max(a.get("clicks", 0), 1), 2),
    } for a in top_earners]

    # Projected revenue (based on active paid ads)
    active_paid = [a for a in paid_ads if a.get("is_active")]
    projected_monthly = sum(a.get("ad_value", 0) for a in active_paid)
    projected_yearly = projected_monthly * 12

    return {
        "summary": {
            "total_revenue": total_revenue,
            "paid_ads_count": len(paid_ads),
            "gift_ads_count": len(gift_ads),
            "avg_ad_value": avg_ad_value,
            "cost_per_click": cost_per_click,
            "cost_per_view": cost_per_view,
            "current_month_revenue": current_rev,
            "previous_month_revenue": prev_rev,
            "growth_percent": growth,
            "projected_monthly": projected_monthly,
            "projected_yearly": projected_yearly,
        },
        "position_revenue": position_revenue,
        "monthly_chart": monthly_chart,
        "top_earners": top_earners_clean,
        "breakdown": {
            "paid_revenue": total_revenue,
            "gift_value": sum(a.get("ad_value", 0) for a in gift_ads),
            "active_revenue": sum(a.get("ad_value", 0) for a in active_paid),
            "inactive_revenue": total_revenue - sum(a.get("ad_value", 0) for a in active_paid),
        }
    }


@router.get("/ads/analytics/compare")
async def compare_ad_periods(
    period1_start: str = "", period1_end: str = "",
    period2_start: str = "", period2_end: str = "",
    current_user: dict = Depends(require_super_admin)
):
    """Compare ad performance between two time periods"""
    db = get_db()
    now = datetime.now(timezone.utc)

    # Default: this month vs last month
    if not period1_start:
        first_of_month = now.replace(day=1)
        period1_start = first_of_month.strftime("%Y-%m-%d")
        period1_end = now.strftime("%Y-%m-%d")
        last_month_end = first_of_month - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)
        period2_start = last_month_start.strftime("%Y-%m-%d")
        period2_end = last_month_end.strftime("%Y-%m-%d")

    ads = await db.internal_ads.find({}, {"_id": 0}).to_list(500)
    events = await db.ad_events.find({}, {"_id": 0}).to_list(20000)

    def calc_period(start, end):
        p_events = [e for e in events if start <= e.get("timestamp", "")[:10] <= end]
        p_clicks = len(p_events)
        # Ads created in this period
        p_ads = [a for a in ads if start <= a.get("created_at", "")[:10] <= end]
        p_revenue = sum(a.get("ad_value", 0) for a in p_ads if not a.get("is_gift"))
        p_new_ads = len(p_ads)
        # All active ads during period
        active_in_period = [a for a in ads if
            (not a.get("start_date") or a.get("start_date", "") <= end) and
            (not a.get("end_date") or a.get("end_date", "") >= start)]
        p_views = sum(a.get("views", 0) for a in active_in_period)
        p_ctr = round((p_clicks / max(p_views, 1)) * 100, 2)
        return {
            "start": start, "end": end,
            "clicks": p_clicks, "views": p_views, "ctr": p_ctr,
            "revenue": p_revenue, "new_ads": p_new_ads,
            "active_ads": len(active_in_period),
        }

    p1 = calc_period(period1_start, period1_end)
    p2 = calc_period(period2_start, period2_end)

    # Calculate changes
    def pct_change(new, old):
        if old == 0:
            return 100.0 if new > 0 else 0
        return round(((new - old) / old) * 100, 1)

    changes = {
        "clicks": pct_change(p1["clicks"], p2["clicks"]),
        "views": pct_change(p1["views"], p2["views"]),
        "ctr": round(p1["ctr"] - p2["ctr"], 2),
        "revenue": pct_change(p1["revenue"], p2["revenue"]),
        "new_ads": pct_change(p1["new_ads"], p2["new_ads"]),
    }

    return {"period1": p1, "period2": p2, "changes": changes}


@router.get("/ads/analytics/export")
async def export_ad_analytics(format: str = "excel", current_user: dict = Depends(require_super_admin)):
    """Export ad analytics as Excel or CSV"""
    from fastapi.responses import StreamingResponse
    import io

    db = get_db()
    ads = await db.internal_ads.find({}, {"_id": 0}).to_list(500)

    pos_labels = {"banner": "بانر", "sidebar": "جانبي", "inline": "داخلي", "dashboard": "لوحة التحكم"}

    if format == "csv":
        output = io.StringIO()
        output.write('\ufeff')  # BOM for Arabic support
        output.write("العنوان,الموقع,المشاهدات,النقرات,CTR%,القيمة,هدية,الحالة,تاريخ الإنشاء\n")
        for a in ads:
            views = a.get("views", 0)
            clicks = a.get("clicks", 0)
            ctr = round((clicks / max(views, 1)) * 100, 2)
            pos = pos_labels.get(a.get("position", ""), a.get("position", ""))
            gift = "نعم" if a.get("is_gift") else "لا"
            status = "نشط" if a.get("is_active") else "متوقف"
            output.write(f'"{a.get("title","")}","{pos}",{views},{clicks},{ctr},{a.get("ad_value",0)},{gift},{status},"{a.get("created_at","")[:10]}"\n')

        # Summary
        total_rev = sum(a.get("ad_value", 0) for a in ads if not a.get("is_gift"))
        total_views = sum(a.get("views", 0) for a in ads)
        total_clicks = sum(a.get("clicks", 0) for a in ads)
        output.write(f'\n"الملخص",,{total_views},{total_clicks},{round((total_clicks/max(total_views,1))*100,2)},{total_rev},,,""\n')

        content = output.getvalue().encode('utf-8-sig')
        return StreamingResponse(
            io.BytesIO(content),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=ad_analytics_report.csv"}
        )

    # Excel format
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # Sheet 1: All Ads
    ws = wb.active
    ws.title = "الإعلانات"
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["العنوان", "الموقع", "المشاهدات", "النقرات", "CTR%", "القيمة (ج.م)", "هدية", "الحالة", "تاريخ الإنشاء"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, a in enumerate(ads, 2):
        views = a.get("views", 0)
        clicks = a.get("clicks", 0)
        ctr = round((clicks / max(views, 1)) * 100, 2)
        row_data = [
            a.get("title", ""),
            pos_labels.get(a.get("position", ""), a.get("position", "")),
            views, clicks, ctr,
            a.get("ad_value", 0),
            "نعم" if a.get("is_gift") else "لا",
            "نشط" if a.get("is_active") else "متوقف",
            a.get("created_at", "")[:10]
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Color status
        status_cell = ws.cell(row=i, column=8)
        status_cell.fill = green_fill if a.get("is_active") else red_fill

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'A'].width = 18

    # Sheet 2: Financial Summary
    ws2 = wb.create_sheet("الملخص المالي")
    ws2.sheet_view.rightToLeft = True

    paid_ads = [a for a in ads if not a.get("is_gift")]
    total_revenue = sum(a.get("ad_value", 0) for a in paid_ads)
    total_clicks_all = sum(a.get("clicks", 0) for a in ads)
    total_views_all = sum(a.get("views", 0) for a in ads)

    summary_data = [
        ("إجمالي الإعلانات", len(ads)),
        ("إعلانات مدفوعة", len(paid_ads)),
        ("إعلانات هدية", len(ads) - len(paid_ads)),
        ("إعلانات نشطة", len([a for a in ads if a.get("is_active")])),
        ("إجمالي الإيرادات (ج.م)", total_revenue),
        ("متوسط قيمة الإعلان (ج.م)", round(total_revenue / max(len(paid_ads), 1), 2)),
        ("إجمالي المشاهدات", total_views_all),
        ("إجمالي النقرات", total_clicks_all),
        ("متوسط CTR%", round((total_clicks_all / max(total_views_all, 1)) * 100, 2)),
        ("تكلفة النقرة CPC (ج.م)", round(total_revenue / max(total_clicks_all, 1), 2)),
        ("الإيرادات المتوقعة سنوياً (ج.م)", sum(a.get("ad_value", 0) for a in paid_ads if a.get("is_active")) * 12),
    ]

    ws2.cell(row=1, column=1, value="المؤشر").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=1).border = thin_border
    ws2.cell(row=1, column=2, value="القيمة").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=2).border = thin_border

    for i, (label, val) in enumerate(summary_data, 2):
        ws2.cell(row=i, column=1, value=label).border = thin_border
        cell = ws2.cell(row=i, column=2, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 20

    # Sheet 3: Revenue by Position
    ws3 = wb.create_sheet("الإيرادات حسب الموقع")
    ws3.sheet_view.rightToLeft = True
    pos_headers = ["الموقع", "عدد الإعلانات", "الإيرادات (ج.م)", "المشاهدات", "النقرات", "CPC (ج.م)"]
    for col, h in enumerate(pos_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    pos_stats = {}
    for a in paid_ads:
        pos = a.get("position", "other")
        pos_stats.setdefault(pos, {"count": 0, "revenue": 0, "views": 0, "clicks": 0})
        pos_stats[pos]["count"] += 1
        pos_stats[pos]["revenue"] += a.get("ad_value", 0)
        pos_stats[pos]["views"] += a.get("views", 0)
        pos_stats[pos]["clicks"] += a.get("clicks", 0)

    for i, (pos, s) in enumerate(pos_stats.items(), 2):
        row = [pos_labels.get(pos, pos), s["count"], s["revenue"], s["views"], s["clicks"],
               round(s["revenue"] / max(s["clicks"], 1), 2)]
        for col, val in enumerate(row, 1):
            cell = ws3.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    for col in range(1, len(pos_headers) + 1):
        ws3.column_dimensions[chr(64 + col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ad_analytics_report.xlsx"}
    )


@router.post("/ads/analytics/send-weekly-report")
async def send_weekly_ad_report(current_user: dict = Depends(require_super_admin)):
    """Send weekly ad performance report via email"""
    from email_service import EmailService

    db = get_db()
    email_svc = EmailService()
    now = datetime.now(timezone.utc)
    week_ago = (now - timedelta(days=7)).isoformat()

    ads = await db.internal_ads.find({}, {"_id": 0}).to_list(500)
    events = await db.ad_events.find({"timestamp": {"$gte": week_ago}}, {"_id": 0}).to_list(10000)

    total_ads = len(ads)
    active_ads = len([a for a in ads if a.get("is_active")])
    total_views = sum(a.get("views", 0) for a in ads)
    total_clicks = sum(a.get("clicks", 0) for a in ads)
    week_clicks = len(events)
    total_revenue = sum(a.get("ad_value", 0) for a in ads if not a.get("is_gift"))
    avg_ctr = round((total_clicks / max(total_views, 1)) * 100, 2)

    # Top 5 ads by CTR
    top_ads = sorted(
        [a for a in ads if a.get("views", 0) > 0],
        key=lambda x: (x.get("clicks", 0) / max(x.get("views", 0), 1)),
        reverse=True
    )[:5]

    top_ads_html = ""
    for a in top_ads:
        ctr = round((a.get("clicks", 0) / max(a.get("views", 0), 1)) * 100, 2)
        top_ads_html += f"""
        <tr>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;">{a.get('title','')}</td>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;">{a.get('views',0):,}</td>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;">{a.get('clicks',0):,}</td>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;">{ctr}%</td>
        </tr>"""

    html = f"""
    <div dir="rtl" style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;border:1px solid #e5e7eb;">
        <div style="background:linear-gradient(135deg,#111827,#1f2937);padding:24px;text-align:center;">
            <h1 style="color:#fff;margin:0;font-size:20px;">تقرير أداء الإعلانات الأسبوعي</h1>
            <p style="color:#9ca3af;margin:8px 0 0;font-size:13px;">HomeMe - {now.strftime('%Y-%m-%d')}</p>
        </div>
        <div style="padding:24px;">
            <div style="display:flex;flex-wrap:wrap;gap:12px;margin-bottom:24px;">
                <div style="flex:1;min-width:120px;background:#ecfdf5;border-radius:10px;padding:16px;text-align:center;">
                    <div style="font-size:24px;font-weight:900;color:#059669;">{total_revenue:,.0f}</div>
                    <div style="font-size:11px;color:#6b7280;margin-top:4px;">إيرادات (ج.م)</div>
                </div>
                <div style="flex:1;min-width:120px;background:#eff6ff;border-radius:10px;padding:16px;text-align:center;">
                    <div style="font-size:24px;font-weight:900;color:#2563eb;">{total_views:,}</div>
                    <div style="font-size:11px;color:#6b7280;margin-top:4px;">مشاهدات</div>
                </div>
                <div style="flex:1;min-width:120px;background:#fef3c7;border-radius:10px;padding:16px;text-align:center;">
                    <div style="font-size:24px;font-weight:900;color:#d97706;">{total_clicks:,}</div>
                    <div style="font-size:11px;color:#6b7280;margin-top:4px;">نقرات</div>
                </div>
            </div>
            <div style="background:#f9fafb;border-radius:10px;padding:16px;margin-bottom:24px;">
                <div style="display:flex;justify-content:space-between;margin-bottom:8px;">
                    <span style="color:#6b7280;font-size:13px;">إجمالي الإعلانات</span>
                    <span style="font-weight:700;">{total_ads}</span>
                </div>
                <div style="display:flex;justify-content:space-between;margin-bottom:8px;">
                    <span style="color:#6b7280;font-size:13px;">إعلانات نشطة</span>
                    <span style="font-weight:700;color:#059669;">{active_ads}</span>
                </div>
                <div style="display:flex;justify-content:space-between;margin-bottom:8px;">
                    <span style="color:#6b7280;font-size:13px;">نقرات هذا الأسبوع</span>
                    <span style="font-weight:700;color:#d97706;">{week_clicks}</span>
                </div>
                <div style="display:flex;justify-content:space-between;">
                    <span style="color:#6b7280;font-size:13px;">متوسط CTR</span>
                    <span style="font-weight:700;color:#e11d48;">{avg_ctr}%</span>
                </div>
            </div>
            <h3 style="font-size:15px;margin:0 0 12px;">أفضل الإعلانات أداءً</h3>
            <table style="width:100%;border-collapse:collapse;font-size:13px;">
                <thead>
                    <tr style="background:#f3f4f6;">
                        <th style="padding:8px 12px;text-align:right;">الإعلان</th>
                        <th style="padding:8px 12px;text-align:center;">مشاهدات</th>
                        <th style="padding:8px 12px;text-align:center;">نقرات</th>
                        <th style="padding:8px 12px;text-align:center;">CTR</th>
                    </tr>
                </thead>
                <tbody>{top_ads_html}</tbody>
            </table>
        </div>
        <div style="background:#f9fafb;padding:16px;text-align:center;font-size:11px;color:#9ca3af;">
            هذا تقرير تلقائي من منصة HomeMe - لإيقاف التقارير، تواصل مع الدعم
        </div>
    </div>
    """

    # Send to owner email
    owner = await db.users.find_one({"role": "app_owner"}, {"_id": 0})
    to_email = owner.get("email", current_user.get("email", "")) if owner else current_user.get("email", "")

    if not to_email:
        raise HTTPException(status_code=400, detail="لم يتم العثور على بريد إلكتروني للمالك")

    await email_svc.send_email(to_email, "تقرير أداء الإعلانات الأسبوعي - HomeMe", html)

    return {
        "message": f"تم إرسال التقرير الأسبوعي إلى {to_email}",
        "to_email": to_email,
        "summary": {
            "total_revenue": total_revenue,
            "total_views": total_views,
            "total_clicks": total_clicks,
            "week_clicks": week_clicks,
            "avg_ctr": avg_ctr,
        }
    }


async def send_weekly_report_auto():
    """Auto-send weekly ad report (called by scheduler, no auth needed)"""
    from email_service import EmailService

    db = get_db()
    email_svc = EmailService()
    now = datetime.now(timezone.utc)
    week_ago = (now - timedelta(days=7)).isoformat()

    ads = await db.internal_ads.find({}, {"_id": 0}).to_list(500)
    events = await db.ad_events.find({"timestamp": {"$gte": week_ago}}, {"_id": 0}).to_list(10000)

    total_ads = len(ads)
    active_ads = len([a for a in ads if a.get("is_active")])
    total_views = sum(a.get("views", 0) for a in ads)
    total_clicks = sum(a.get("clicks", 0) for a in ads)
    week_clicks = len(events)
    total_revenue = sum(a.get("ad_value", 0) for a in ads if not a.get("is_gift"))
    avg_ctr = round((total_clicks / max(total_views, 1)) * 100, 2)

    top_ads = sorted(
        [a for a in ads if a.get("views", 0) > 0],
        key=lambda x: (x.get("clicks", 0) / max(x.get("views", 0), 1)),
        reverse=True
    )[:5]

    top_ads_html = ""
    for a in top_ads:
        ctr = round((a.get("clicks", 0) / max(a.get("views", 0), 1)) * 100, 2)
        top_ads_html += f"""
        <tr>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;">{a.get('title','')}</td>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;">{a.get('views',0):,}</td>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;">{a.get('clicks',0):,}</td>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;">{ctr}%</td>
        </tr>"""

    html = f"""
    <div dir="rtl" style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;border:1px solid #e5e7eb;">
        <div style="background:linear-gradient(135deg,#111827,#1f2937);padding:24px;text-align:center;">
            <h1 style="color:#fff;margin:0;font-size:20px;">تقرير أداء الإعلانات الأسبوعي (تلقائي)</h1>
            <p style="color:#9ca3af;margin:8px 0 0;font-size:13px;">HomeMe - {now.strftime('%Y-%m-%d')}</p>
        </div>
        <div style="padding:24px;">
            <div style="display:flex;flex-wrap:wrap;gap:12px;margin-bottom:24px;">
                <div style="flex:1;min-width:120px;background:#ecfdf5;border-radius:10px;padding:16px;text-align:center;">
                    <div style="font-size:24px;font-weight:900;color:#059669;">{total_revenue:,.0f}</div>
                    <div style="font-size:11px;color:#6b7280;margin-top:4px;">إيرادات (ج.م)</div>
                </div>
                <div style="flex:1;min-width:120px;background:#eff6ff;border-radius:10px;padding:16px;text-align:center;">
                    <div style="font-size:24px;font-weight:900;color:#2563eb;">{total_views:,}</div>
                    <div style="font-size:11px;color:#6b7280;margin-top:4px;">مشاهدات</div>
                </div>
                <div style="flex:1;min-width:120px;background:#fef3c7;border-radius:10px;padding:16px;text-align:center;">
                    <div style="font-size:24px;font-weight:900;color:#d97706;">{total_clicks:,}</div>
                    <div style="font-size:11px;color:#6b7280;margin-top:4px;">نقرات</div>
                </div>
            </div>
            <div style="background:#f9fafb;border-radius:10px;padding:16px;margin-bottom:24px;">
                <div style="display:flex;justify-content:space-between;margin-bottom:8px;">
                    <span style="color:#6b7280;font-size:13px;">إجمالي الإعلانات</span>
                    <span style="font-weight:700;">{total_ads}</span>
                </div>
                <div style="display:flex;justify-content:space-between;margin-bottom:8px;">
                    <span style="color:#6b7280;font-size:13px;">إعلانات نشطة</span>
                    <span style="font-weight:700;color:#059669;">{active_ads}</span>
                </div>
                <div style="display:flex;justify-content:space-between;margin-bottom:8px;">
                    <span style="color:#6b7280;font-size:13px;">نقرات هذا الأسبوع</span>
                    <span style="font-weight:700;color:#d97706;">{week_clicks}</span>
                </div>
                <div style="display:flex;justify-content:space-between;">
                    <span style="color:#6b7280;font-size:13px;">متوسط CTR</span>
                    <span style="font-weight:700;color:#e11d48;">{avg_ctr}%</span>
                </div>
            </div>
            <h3 style="font-size:15px;margin:0 0 12px;">أفضل الإعلانات أداءً</h3>
            <table style="width:100%;border-collapse:collapse;font-size:13px;">
                <thead>
                    <tr style="background:#f3f4f6;">
                        <th style="padding:8px 12px;text-align:right;">الإعلان</th>
                        <th style="padding:8px 12px;text-align:center;">مشاهدات</th>
                        <th style="padding:8px 12px;text-align:center;">نقرات</th>
                        <th style="padding:8px 12px;text-align:center;">CTR</th>
                    </tr>
                </thead>
                <tbody>{top_ads_html}</tbody>
            </table>
        </div>
        <div style="background:#f9fafb;padding:16px;text-align:center;font-size:11px;color:#9ca3af;">
            هذا تقرير تلقائي أسبوعي من منصة HomeMe - يُرسل كل يوم أحد الساعة 8 صباحاً
        </div>
    </div>
    """

    owner = await db.users.find_one({"role": "app_owner"}, {"_id": 0})
    to_email = owner.get("email", "") if owner else ""

    if not to_email:
        return "No owner email found"

    await email_svc.send_email(to_email, "تقرير أداء الإعلانات الأسبوعي (تلقائي) - HomeMe", html)
    return f"Weekly ad report sent to {to_email}"


@router.get("/ads/analytics/export-pdf")
async def export_ad_analytics_pdf(current_user: dict = Depends(require_super_admin)):
    """Export ad analytics as PDF with Arabic support.
    
    Performance: PDF rendering (CPU-bound) moved off the event loop via asyncio.to_thread.
    arabic_reshaper + reportlab were blocking ~5-10s per request on production.
    """
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import arabic_reshaper
    from bidi.algorithm import get_display
    import io

    db = get_db()
    ads = await db.internal_ads.find({}, {"_id": 0}).to_list(500)
    now = datetime.now(timezone.utc)

    def _build_pdf(ads_data, now_dt):
        """CPU-bound PDF builder — runs in threadpool."""
        def ar(text):
            try:
                reshaped = arabic_reshaper.reshape(str(text))
                return get_display(reshaped)
            except Exception:
                return str(text)

        pos_labels = {"banner": ar("بانر"), "sidebar": ar("جانبي"), "inline": ar("داخلي"), "dashboard": ar("لوحة التحكم")}

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title_AR', parent=styles['Title'], fontSize=22, spaceAfter=10, alignment=TA_CENTER)
        subtitle_style = ParagraphStyle('Sub_AR', parent=styles['Normal'], fontSize=10, textColor=colors.gray, alignment=TA_CENTER, spaceAfter=20)
        heading_style = ParagraphStyle('Heading_AR', parent=styles['Heading2'], fontSize=14, spaceAfter=8, spaceBefore=15, textColor=colors.HexColor('#1F2937'))

        elements = []

        # Title
        elements.append(Paragraph(ar("تقرير تحليلات الإعلانات - HomeMe"), title_style))
        elements.append(Paragraph(f"{ar('تاريخ التقرير')}: {now_dt.strftime('%Y-%m-%d %H:%M')} UTC", subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#E5E7EB')))
        elements.append(Spacer(1, 10))

        # Summary
        paid_ads = [a for a in ads_data if not a.get("is_gift")]
        gift_ads = [a for a in ads_data if a.get("is_gift")]
        total_revenue = sum(a.get("ad_value", 0) for a in paid_ads)
        total_views = sum(a.get("views", 0) for a in ads_data)
        total_clicks = sum(a.get("clicks", 0) for a in ads_data)
        avg_ctr = round((total_clicks / max(total_views, 1)) * 100, 2)

        elements.append(Paragraph(ar("الملخص المالي"), heading_style))

        summary_data = [
            [ar("القيمة"), ar("المؤشر")],
            [str(len(ads_data)), ar("إجمالي الإعلانات")],
            [str(len([a for a in ads_data if a.get("is_active")])), ar("إعلانات نشطة")],
            [str(len(paid_ads)), ar("إعلانات مدفوعة")],
            [str(len(gift_ads)), ar("إعلانات هدية")],
            [f"{total_revenue:,.2f}", ar("إجمالي الإيرادات (ج.م)")],
            [f"{round(total_revenue / max(len(paid_ads), 1), 2):,.2f}", ar("متوسط قيمة الإعلان (ج.م)")],
            [f"{total_views:,}", ar("إجمالي المشاهدات")],
            [f"{total_clicks:,}", ar("إجمالي النقرات")],
            [f"{avg_ctr}%", ar("متوسط نسبة النقر CTR")],
            [f"{round(total_revenue / max(total_clicks, 1), 2):,.2f}", ar("تكلفة النقرة CPC (ج.م)")],
            [f"{sum(a.get('ad_value', 0) for a in paid_ads if a.get('is_active')) * 12:,.2f}", ar("الإيرادات المتوقعة سنوياً (ج.م)")],
        ]

        t_summary = Table(summary_data, colWidths=[150, 250])
        t_summary.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(t_summary)
        elements.append(Spacer(1, 15))

        # Revenue by Position
        elements.append(Paragraph(ar("الإيرادات حسب الموقع"), heading_style))
        pos_stats = {}
        for a in paid_ads:
            pos = a.get("position", "other")
            pos_stats.setdefault(pos, {"count": 0, "revenue": 0, "views": 0, "clicks": 0})
            pos_stats[pos]["count"] += 1
            pos_stats[pos]["revenue"] += a.get("ad_value", 0)
            pos_stats[pos]["views"] += a.get("views", 0)
            pos_stats[pos]["clicks"] += a.get("clicks", 0)

        pos_data = [[ar("CPC"), ar("النقرات"), ar("المشاهدات"), ar("الإيرادات"), ar("العدد"), ar("الموقع")]]
        for pos, s in pos_stats.items():
            pos_data.append([
                f"{round(s['revenue'] / max(s['clicks'], 1), 2):,.2f}",
                f"{s['clicks']:,}", f"{s['views']:,}",
                f"{s['revenue']:,.2f}", str(s["count"]),
                pos_labels.get(pos, ar(pos))
            ])

        t_pos = Table(pos_data, colWidths=[70, 70, 80, 90, 50, 100])
        t_pos.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EFF6FF')]),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(t_pos)
        elements.append(Spacer(1, 15))

        # All Ads
        elements.append(Paragraph(ar("أداء جميع الإعلانات"), heading_style))
        ads_rows = [[ar("الحالة"), ar("القيمة"), "CTR%", ar("النقرات"), ar("المشاهدات"), ar("الموقع"), ar("العنوان")]]
        for a in sorted(ads_data, key=lambda x: x.get("ad_value", 0), reverse=True):
            views = a.get("views", 0)
            clicks = a.get("clicks", 0)
            ctr = round((clicks / max(views, 1)) * 100, 2)
            title_text = a.get("title", "")[:25]
            try:
                title_text = ar(title_text)
            except Exception:
                pass
            ads_rows.append([
                ar("نشط") if a.get("is_active") else ar("متوقف"),
                ar("هدية") if a.get("is_gift") else f"{a.get('ad_value', 0):,.0f}",
                f"{ctr}%", f"{clicks:,}", f"{views:,}",
                pos_labels.get(a.get("position", ""), a.get("position", "")),
                title_text
            ])

        t_ads = Table(ads_rows, colWidths=[50, 60, 45, 50, 60, 70, 130])
        t_ads.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ECFDF5')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(t_ads)

        # Footer
        elements.append(Spacer(1, 20))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#D1D5DB')))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray, alignment=TA_CENTER)
        elements.append(Paragraph(f"HomeMe Platform - {ar('تقرير تلقائي')} - {now_dt.strftime('%Y-%m-%d')}", footer_style))

        doc.build(elements)
        buf.seek(0)
        return buf

    # Off-load the entire PDF rendering to a worker thread so it doesn't block the event loop
    buffer = await asyncio.to_thread(_build_pdf, ads, now)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=ad_analytics_report.pdf"}
    )


async def check_ctr_alerts_and_notify():
    """Check ads for high CTR and send push notifications to admins/owner"""
    db = get_db()
    ads = await db.internal_ads.find({"is_active": True}, {"_id": 0}).to_list(500)

    high_ctr_ads = []
    for a in ads:
        views = a.get("views", 0)
        clicks = a.get("clicks", 0)
        if views < 10:
            continue
        ctr = round((clicks / views) * 100, 2)
        if ctr >= 5:
            # Check if we already notified about this ad recently
            existing = await db.ctr_alerts.find_one({
                "ad_id": a.get("id"),
                "created_at": {"$gte": (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()}
            })
            if not existing:
                high_ctr_ads.append({"ad": a, "ctr": ctr})

    if not high_ctr_ads:
        return 0

    # Get admin/owner push subscriptions
    admin_subs = await db.push_subscriptions.find({
        "is_active": True,
        "role": {"$in": ["app_owner", "super_admin"]}
    }).to_list(100)

    # Also get by user role from users collection
    if not admin_subs:
        admins = await db.users.find(
            {"role": {"$in": ["app_owner", "super_admin"]}},
            {"_id": 0, "id": 1}
        ).to_list(10)
        admin_ids = [u["id"] for u in admins]
        admin_subs = await db.push_subscriptions.find({
            "is_active": True,
            "user_id": {"$in": admin_ids}
        }).to_list(100)

    sent_count = 0
    for item in high_ctr_ads:
        ad = item["ad"]
        ctr = item["ctr"]

        # Store alert record
        await db.ctr_alerts.insert_one({
            "ad_id": ad.get("id"),
            "ad_title": ad.get("title"),
            "ctr": ctr,
            "views": ad.get("views", 0),
            "clicks": ad.get("clicks", 0),
            "created_at": datetime.now(timezone.utc).isoformat()
        })

        # Also store as notification
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "type": "ctr_alert",
            "title": f"CTR عالي: {ad.get('title', '')}",
            "message": f"الإعلان حقق CTR {ctr}% ({ad.get('clicks',0)} نقرة من {ad.get('views',0)} مشاهدة)",
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "data": {"ad_id": ad.get("id"), "ctr": ctr}
        })

        # Send push notifications
        for sub in admin_subs:
            try:
                from push_notification_service import PushNotificationService
                PushNotificationService.send_notification(
                    subscription=sub,
                    title=f"CTR عالي: {ad.get('title', '')}",
                    body=f"الإعلان حقق نسبة نقر {ctr}% - {ad.get('clicks',0)} نقرة",
                    url="/app/ad-analytics?tab=alerts",
                    tag=f"ctr-alert-{ad.get('id')}",
                    require_interaction=True,
                    data={"type": "ctr_alert", "ad_id": ad.get("id"), "ctr": ctr}
                )
                sent_count += 1
            except Exception as e:
                logging.error(f"Failed to send CTR push notification: {e}")

    logging.info(f"CTR alerts: {len(high_ctr_ads)} ads, {sent_count} notifications sent")
    return len(high_ctr_ads)


UPLOAD_DIR = "/app/uploads/ads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


@router.post("/ads/upload-media")
async def upload_ad_media(file: UploadFile = File(...), current_user: dict = Depends(require_super_admin)):
    """Upload image or video for an ad"""
    allowed_image = [".jpg", ".jpeg", ".png", ".gif", ".webp"]
    allowed_video = [".mp4", ".webm", ".mov"]
    allowed = allowed_image + allowed_video

    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in allowed:
        raise HTTPException(status_code=400, detail=f"نوع ملف غير مدعوم. المسموح: {', '.join(allowed)}")

    content = await file.read()
    max_size = 50 * 1024 * 1024  # 50MB
    if len(content) > max_size:
        raise HTTPException(status_code=400, detail="حجم الملف كبير جداً (الحد الأقصى 50 ميجا)")

    filename = f"{uuid.uuid4().hex[:12]}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    with open(filepath, "wb") as f:
        f.write(content)

    # Dual-write to MongoDB so image survives container rebuilds / deployments
    try:
        from services.media_store import save_to_db
        await save_to_db("ads", filename, file.content_type or "", content)
    except Exception as _e:
        import logging as _lg
        _lg.warning(f"ads upload DB backup failed: {_e}")

    media_type = "video" if ext in allowed_video else "image"
    media_url = f"/api/ads/media/{filename}"

    return {"url": media_url, "filename": filename, "type": media_type, "size": len(content)}


@router.get("/ads/media/{filename}")
async def serve_ad_media(filename: str):
    """Serve uploaded ad media with full self-healing (backup → DB → 404)"""
    from fastapi.responses import FileResponse, Response
    filepath = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(filepath):
        # Self-heal: try backup snapshots first
        try:
            from services.media_backup import restore_file
            restore_file("ads", filename)
        except Exception:
            pass
        # Self-heal: try MongoDB persistent store (survives deployments)
        if not os.path.exists(filepath):
            try:
                from services.media_store import load_from_db
                result = await load_from_db("ads", filename)
                if result:
                    content_type, data = result
                    # Write back to disk for cache
                    try:
                        with open(filepath, "wb") as f:
                            f.write(data)
                    except Exception:
                        pass
                    return Response(content=data, media_type=content_type)
            except Exception:
                pass
        if not os.path.exists(filepath):
            raise HTTPException(status_code=404, detail="ملف غير موجود")
    return FileResponse(filepath)


@router.get("/ads/check-url")
async def check_ad_url(url: str, current_user: dict = Depends(get_current_user)):
    """التحقق من صحة الرابط (server-side لتجنّب CORS)"""
    if not url or not url.strip():
        return {"ok": False, "status": 0, "error": "empty"}
    full = url.strip()
    if not full.startswith(("http://", "https://")):
        full = "https://" + full
    try:
        async with httpx.AsyncClient(timeout=8.0, follow_redirects=True) as client:
            try:
                resp = await client.head(full)
                # بعض المواقع لا تدعم HEAD — نجرب GET خفيف
                if resp.status_code >= 400 and resp.status_code != 405:
                    pass
                elif resp.status_code == 405:
                    resp = await client.get(full, headers={"Range": "bytes=0-0"})
                return {
                    "ok": 200 <= resp.status_code < 400,
                    "status": resp.status_code,
                    "final_url": str(resp.url),
                }
            except httpx.HTTPError:
                # محاولة GET كـ fallback
                resp = await client.get(full, headers={"Range": "bytes=0-0"})
                return {
                    "ok": 200 <= resp.status_code < 400,
                    "status": resp.status_code,
                    "final_url": str(resp.url),
                }
    except httpx.TimeoutException:
        return {"ok": False, "status": 0, "error": "timeout"}
    except httpx.ConnectError:
        return {"ok": False, "status": 0, "error": "connection_failed"}
    except Exception as e:
        return {"ok": False, "status": 0, "error": str(e)[:100]}

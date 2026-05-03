"""
Bulk Resident Import — upload Excel/CSV file with hundreds of residents at once.

Flow:
  1. GET /api/residents/bulk-import/template → download blank Excel template.
  2. POST /api/residents/bulk-import/preview (multipart) → parse + validate
     without persisting; returns row-by-row preview {ok, errors, warnings}.
  3. POST /api/residents/bulk-import/commit → actually insert valid rows,
     return final stats {created, skipped, failed}.

Column mapping is fuzzy (Arabic + English aliases). Any extra columns are
ignored. Validation:
  - full_name + unit_number are REQUIRED.
  - email & username uniqueness checked against DB + within the same file.
  - password auto-generated if missing (8-char alphanumeric).
  - username auto-generated from name + unit if missing.
"""
from fastapi import APIRouter, Depends, File, UploadFile, HTTPException, Request
from fastapi.responses import StreamingResponse
from typing import List, Dict, Optional
from datetime import datetime, timezone
import io
import re
import uuid
import secrets
import string
import logging

from openpyxl import Workbook, load_workbook
import bcrypt

from database import get_db
from auth_deps import require_admin

router = APIRouter(prefix="/api")
logger = logging.getLogger(__name__)

# Column aliases — Arabic + English variants the user might put in their sheet
COL_ALIASES = {
    "full_name": ["full_name", "name", "الاسم", "الاسم الكامل", "اسم الساكن", "الاسم بالكامل"],
    "unit_number": ["unit_number", "unit", "apartment", "الوحدة", "رقم الوحدة", "الشقة", "رقم الشقة", "الفيلا", "رقم الفيلا"],
    "phone": ["phone", "mobile", "الجوال", "الهاتف", "رقم الهاتف", "الموبايل", "رقم الموبايل"],
    "email": ["email", "البريد", "البريد الإلكتروني", "الإيميل"],
    "username": ["username", "user", "اسم المستخدم", "اليوزر"],
    "password": ["password", "pwd", "كلمة المرور", "الباسورد"],
    "role": ["role", "الدور", "النوع", "نوع المستخدم"],
    "national_id": ["national_id", "national id", "الرقم القومي", "البطاقة"],
}
ROLE_VALUES = {"resident", "security", "staff", "admin"}
TEMPLATE_HEADERS = [
    "الاسم الكامل", "رقم الوحدة", "رقم الهاتف",
    "البريد الإلكتروني", "اسم المستخدم", "كلمة المرور",
    "الدور (resident/security/staff)", "الرقم القومي"
]


def _normalize(s: str) -> str:
    """Lower-case + strip + remove diacritics for header comparison."""
    return (s or "").strip().lower().replace("ـ", "").replace("ال", "ال")


def _build_header_index(headers: List[str]) -> Dict[str, int]:
    """Map our internal column names to actual Excel column indices."""
    idx_map: Dict[str, int] = {}
    for i, raw in enumerate(headers):
        norm = _normalize(raw)
        for canonical, aliases in COL_ALIASES.items():
            if canonical in idx_map:
                continue
            for alias in aliases:
                if _normalize(alias) == norm:
                    idx_map[canonical] = i
                    break
    return idx_map


def _gen_password(length: int = 8) -> str:
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


def _slugify_username(name: str, unit: str) -> str:
    """Generate a user-friendly username from Arabic/English name + unit."""
    # Strip non-alphanumeric (keep underscores)
    base = re.sub(r"[^\w\u0600-\u06FF]+", "_", (name or "").strip()).strip("_").lower()[:20]
    suffix = re.sub(r"[^\w]+", "", str(unit or "")).lower()[:8]
    short = f"{base}_{suffix}" if suffix else base
    if not short:
        short = f"user_{uuid.uuid4().hex[:6]}"
    return short[:32]


def _parse_workbook(file_bytes: bytes, filename: str) -> List[Dict]:
    """Read xlsx/xls/csv into list of dicts using fuzzy header matching."""
    rows: List[Dict] = []
    if filename.lower().endswith(".csv"):
        import csv
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
    else:
        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            all_rows = [[c.value for c in row] for row in ws.iter_rows()]
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"تعذّر قراءة الملف: {e}")

    if not all_rows or len(all_rows) < 2:
        raise HTTPException(status_code=400, detail="الملف فارغ أو لا يحتوي على بيانات")

    headers = [str(h or "").strip() for h in all_rows[0]]
    idx = _build_header_index(headers)
    if "full_name" not in idx or "unit_number" not in idx:
        raise HTTPException(
            status_code=400,
            detail="الأعمدة المطلوبة 'الاسم الكامل' و 'رقم الوحدة' غير موجودة في الملف",
        )

    for row_no, row in enumerate(all_rows[1:], start=2):
        # Skip fully empty rows
        if not row or all((c is None or str(c).strip() == "") for c in row):
            continue
        rec = {"_row": row_no}
        for k, i in idx.items():
            v = row[i] if i < len(row) else None
            if v is None:
                rec[k] = ""
            elif isinstance(v, (int, float)):
                rec[k] = str(int(v)) if float(v).is_integer() else str(v)
            else:
                rec[k] = str(v).strip()
        rows.append(rec)
    return rows


async def _validate_rows(db, rows: List[Dict], compound_id: str) -> Dict:
    """Return {valid: [...], invalid: [{row, errors:[]}], summary: {...}}."""
    valid = []
    invalid = []
    seen_usernames = set()
    seen_emails = set()

    # Pre-fetch existing usernames + emails from DB for quick membership tests
    existing_users = await db.users.find({}, {"_id": 0, "username": 1, "email": 1}).to_list(50000)
    existing_usernames = {u.get("username", "") for u in existing_users if u.get("username")}
    existing_emails = {u.get("email", "") for u in existing_users if u.get("email")}

    for r in rows:
        errors = []
        warnings = []

        full_name = (r.get("full_name") or "").strip()
        unit_number = (r.get("unit_number") or "").strip()
        if not full_name:
            errors.append("الاسم الكامل مطلوب")
        if not unit_number:
            errors.append("رقم الوحدة مطلوب")

        # Email validation (basic)
        email = (r.get("email") or "").strip()
        if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
            errors.append("البريد الإلكتروني غير صالح")
        if email and email in existing_emails:
            errors.append(f"البريد '{email}' مستخدم مسبقاً في النظام")
        if email and email in seen_emails:
            errors.append("البريد مكرر داخل الملف")

        # Username
        username = (r.get("username") or "").strip()
        if not username:
            username = _slugify_username(full_name, unit_number)
            warnings.append(f"اسم المستخدم سيُولَّد تلقائياً: {username}")
        # Make unique by appending suffix if needed
        original_uname = username
        suffix_n = 0
        while username in existing_usernames or username in seen_usernames:
            suffix_n += 1
            username = f"{original_uname}{suffix_n}"
            warnings.append(f"اسم المستخدم مكرر، تم تعديله إلى: {username}")

        # Password
        password = (r.get("password") or "").strip()
        if not password:
            password = _gen_password()
            warnings.append(f"كلمة المرور سيتم توليدها: {password}")
        elif len(password) < 6:
            errors.append("كلمة المرور يجب أن تكون 6 أحرف على الأقل")

        # Role
        role = (r.get("role") or "resident").strip().lower()
        if role not in ROLE_VALUES:
            role = "resident"
            warnings.append("الدور غير معروف، تم استخدام 'resident'")

        rec = {
            "_row": r["_row"],
            "full_name": full_name,
            "unit_number": unit_number,
            "phone": r.get("phone", ""),
            "email": email,
            "username": username,
            "password": password,
            "role": role,
            "national_id": r.get("national_id", ""),
            "warnings": warnings,
        }
        if errors:
            invalid.append({"row": r["_row"], "data": rec, "errors": errors})
        else:
            seen_usernames.add(username)
            if email:
                seen_emails.add(email)
            valid.append(rec)

    return {
        "valid": valid,
        "invalid": invalid,
        "summary": {
            "total": len(rows),
            "valid": len(valid),
            "invalid": len(invalid),
        },
    }


def _resolve_target_compound(request: Request, current_user: dict) -> str:
    cid = (
        request.headers.get("X-Active-Compound-Id")
        or request.headers.get("x-active-compound-id")
        or current_user.get("compound_id")
    )
    if not cid or cid == "default-compound":
        raise HTTPException(status_code=400, detail="يرجى اختيار كمبوند أولاً (X-Active-Compound-Id)")
    return cid


async def _enforce_tenant(db, current_user: dict, compound_id: str):
    role = current_user.get("role")
    if role in ("app_owner", "super_admin"):
        return
    if role in ("company_admin", "assistant_manager", "accountant"):
        cmpd = await db.compounds.find_one(
            {"id": compound_id}, {"_id": 0, "company_id": 1, "management_company_id": 1}
        )
        cu_company = current_user.get("company_id")
        if not cmpd or cu_company not in (cmpd.get("company_id"), cmpd.get("management_company_id")):
            raise HTTPException(status_code=403, detail="غير مصرح بالعمل خارج شركتك")
        return
    if role == "admin":
        if current_user.get("compound_id") != compound_id:
            raise HTTPException(status_code=403, detail="غير مصرح بالعمل خارج كمبوندك")


@router.get("/residents/bulk-import/template")
async def download_template(current_user: dict = Depends(require_admin)):
    """Download a blank Excel template with proper headers + sample row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "السكان"
    # RTL sheet
    ws.sheet_view.rightToLeft = True
    # Headers
    ws.append(TEMPLATE_HEADERS)
    # Sample row to guide users
    ws.append([
        "أحمد محمد علي", "A-12", "01001234567",
        "ahmed@example.com", "ahmed_a12", "Pass1234!",
        "resident", "29012345678901"
    ])
    ws.append([
        "ساره خالد إبراهيم", "B-7", "01112345678",
        "", "", "", "resident", ""
    ])
    # Set column widths
    for col_letter, width in zip("ABCDEFGH", [25, 12, 16, 28, 18, 16, 28, 18]):
        ws.column_dimensions[col_letter].width = width
    # Bold header row
    from openpyxl.styles import Font, PatternFill, Alignment
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="6366F1")
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = align

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=residents_template.xlsx"},
    )


@router.post("/residents/bulk-import/preview")
async def preview_bulk_import(
    request: Request,
    file: UploadFile = File(...),
    current_user: dict = Depends(require_admin),
):
    """Parse + validate the uploaded file. Returns preview without saving anything."""
    db = get_db()
    target_compound = _resolve_target_compound(request, current_user)
    await _enforce_tenant(db, current_user, target_compound)

    data = await file.read()
    if len(data) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="حجم الملف يجب أن يكون أقل من 5MB")
    rows = _parse_workbook(data, file.filename or "file.xlsx")
    if not rows:
        raise HTTPException(status_code=400, detail="لا توجد صفوف بيانات في الملف")

    result = await _validate_rows(db, rows, target_compound)
    result["compound_id"] = target_compound
    return result


@router.post("/residents/bulk-import/commit")
async def commit_bulk_import(
    request: Request,
    file: UploadFile = File(...),
    current_user: dict = Depends(require_admin),
):
    """Re-parse + persist only the VALID rows. Returns final stats."""
    db = get_db()
    target_compound = _resolve_target_compound(request, current_user)
    await _enforce_tenant(db, current_user, target_compound)

    data = await file.read()
    if len(data) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="حجم الملف يجب أن يكون أقل من 5MB")
    rows = _parse_workbook(data, file.filename or "file.xlsx")
    val = await _validate_rows(db, rows, target_compound)

    created = 0
    failed = 0
    failed_rows: List[Dict] = []
    credentials: List[Dict] = []
    now = datetime.now(timezone.utc)

    for rec in val["valid"]:
        try:
            password_hash = bcrypt.hashpw(rec["password"].encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
            user_doc = {
                "id": str(uuid.uuid4()),
                "username": rec["username"],
                "email": rec.get("email") or None,
                "password_hash": password_hash,
                "role": rec["role"],
                "compound_id": target_compound,
                "family_id": None,
                "full_name": rec["full_name"],
                "phone": rec.get("phone") or None,
                "unit_number": rec["unit_number"],
                "national_id": rec.get("national_id") or None,
                "is_family_head": True,
                "is_active": True,
                "created_at": now,
                "created_via": "bulk_import",
                "imported_by": current_user.get("id"),
                "profile_picture_url": None,
            }
            await db.users.insert_one(user_doc)
            credentials.append({
                "full_name": rec["full_name"],
                "username": rec["username"],
                "password": rec["password"],
                "unit_number": rec["unit_number"],
            })
            created += 1
        except Exception as e:
            failed += 1
            failed_rows.append({"row": rec["_row"], "name": rec.get("full_name", ""), "error": str(e)})
            logger.error(f"Bulk import row {rec.get('_row')} failed: {e}")

    return {
        "created": created,
        "failed": failed,
        "skipped": len(val["invalid"]),
        "total": val["summary"]["total"],
        "failed_rows": failed_rows,
        "credentials": credentials,  # one-time access for admin to share with residents
    }

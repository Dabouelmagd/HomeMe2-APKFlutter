"""
Excel/PDF Export & Resident Profile routes
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import Response
from datetime import datetime, timezone, timedelta
from typing import Optional
import io
import logging

from database import get_db
from auth_deps import get_current_user, require_admin
from helpers import serialize_datetime
from pdf_report_service import PDFReportService
from plan_limits import gate_company_feature

router = APIRouter(prefix="/api")


@router.get("/financial/export-excel")
async def export_financial_excel(
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: dict = Depends(require_admin)
):
    await gate_company_feature(current_user, "pdf_excel_exports", "تصدير Excel/PDF")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    db = get_db()
    try:
        compound_id = current_user["compound_id"]
        current_year = year or datetime.now().year

        wb = openpyxl.Workbook()
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        green_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
        red_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def style_header(ws, row, cols):
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        # Sheet 1: Balance Sheet
        ws1 = wb.active
        ws1.title = "Balance Sheet"
        ws1.sheet_properties.sheetView = openpyxl.worksheet.views.SheetView(rightToLeft=True)

        expenses = await db.expenses.find({"compound_id": compound_id}, {"_id": 0}).to_list(500)
        revenues = await db.revenue.find({"compound_id": compound_id}, {"_id": 0}).to_list(500)
        total_exp = sum(float(e.get("amount", 0)) for e in expenses)
        total_rev = sum(float(r.get("amount", 0)) for r in revenues)

        ws1.append(["الميزانية العمومية / Balance Sheet", "", "", ""])
        ws1.merge_cells('A1:D1')
        ws1.cell(1, 1).font = Font(bold=True, size=14)
        ws1.append([])
        ws1.append(["البند", "المبلغ"])
        style_header(ws1, 3, 2)
        ws1.append(["إجمالي الإيرادات", total_rev])
        ws1.append(["إجمالي المصروفات", total_exp])
        ws1.append(["صافي الرصيد", total_rev - total_exp])
        ws1.cell(6, 2).font = Font(bold=True, color="FF0000" if total_rev - total_exp < 0 else "008000")

        ws1.append([])
        ws1.append(["المصروفات حسب التصنيف", "المبلغ"])
        style_header(ws1, 8, 2)
        exp_by_cat = {}
        for e in expenses:
            cat = e.get("category", "other")
            exp_by_cat[cat] = exp_by_cat.get(cat, 0) + float(e.get("amount", 0))
        for cat, amt in exp_by_cat.items():
            ws1.append([cat, amt])
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 20

        # Sheet 2: Expenses Detail
        ws2 = wb.create_sheet("Expenses")
        ws2.sheet_properties.sheetView = openpyxl.worksheet.views.SheetView(rightToLeft=True)
        ws2.append(["الوصف", "التصنيف", "المبلغ", "التاريخ", "الحالة"])
        style_header(ws2, 1, 5)
        for e in expenses:
            ws2.append([e.get("description", ""), e.get("category", ""), float(e.get("amount", 0)), str(e.get("date", ""))[:10], e.get("status", "")])
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws2.column_dimensions[col].width = 22

        # Sheet 3: Unit Charges
        ws3 = wb.create_sheet("Unit Charges")
        ws3.sheet_properties.sheetView = openpyxl.worksheet.views.SheetView(rightToLeft=True)
        query = {"compound_id": compound_id}
        if month:
            query["month"] = month
        if year:
            query["year"] = year
        charges = await db.unit_charges.find(query, {"_id": 0}).sort("unit_number", 1).to_list(500)
        ws3.append(["الوحدة", "المقيم", "الالتزام", "المبلغ", "الحالة", "تاريخ السداد", "الشهر", "السنة"])
        style_header(ws3, 1, 8)
        for c in charges:
            paid_at = ""
            if c.get("paid_at"):
                pa = c["paid_at"]
                paid_at = pa.strftime('%Y-%m-%d') if hasattr(pa, 'strftime') else str(pa)[:10]
            row_num = ws3.max_row + 1
            ws3.append([c.get("unit_number", ""), c.get("resident_name", ""), c.get("title", ""), float(c.get("amount", 0)), "سدد" if c.get("status") == "paid" else "لم يسدد", paid_at, c.get("month", ""), c.get("year", "")])
            fill = green_fill if c.get("status") == "paid" else red_fill
            for col in range(1, 9):
                ws3.cell(row=row_num, column=col).fill = fill
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws3.column_dimensions[col].width = 18

        # Sheet 4: Obligations
        ws4 = wb.create_sheet("Obligations")
        ws4.sheet_properties.sheetView = openpyxl.worksheet.views.SheetView(rightToLeft=True)
        obs = await db.obligations.find({"compound_id": compound_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
        ws4.append(["العنوان", "التصنيف", "المبلغ الإجمالي", "طريقة التوزيع", "عدد الوحدات", "الشهر", "السنة"])
        style_header(ws4, 1, 7)
        for ob in obs:
            ws4.append([ob.get("title", ""), ob.get("category", ""), float(ob.get("total_amount", 0)), ob.get("distribution_label", "بالتساوي"), ob.get("unit_count", 0), ob.get("month", ""), ob.get("year", "")])
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws4.column_dimensions[col].width = 20

        # Sheet 5: Revenue
        ws5 = wb.create_sheet("Revenue")
        ws5.sheet_properties.sheetView = openpyxl.worksheet.views.SheetView(rightToLeft=True)
        ws5.append(["الوصف", "المصدر", "المبلغ", "التاريخ", "الحالة"])
        style_header(ws5, 1, 5)
        for r in revenues:
            ws5.append([r.get("description", ""), r.get("source", ""), float(r.get("amount", 0)), str(r.get("date", ""))[:10], r.get("status", "")])
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws5.column_dimensions[col].width = 22

        buffer = io.BytesIO()
        wb.save(buffer)
        excel_content = buffer.getvalue()
        buffer.close()

        filename = f"financial_report_{current_year}.xlsx"
        return Response(
            content=excel_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logging.error(f"Error exporting Excel: {e}")
        raise HTTPException(status_code=500, detail="Failed to export Excel")


@router.get("/residents/{resident_id}/profile")
async def get_resident_profile(
    resident_id: str,
    sort_order: str = "desc",
    current_user: dict = Depends(get_current_user)
):
    db = get_db()
    try:
        resident = await db.users.find_one({"id": resident_id, "compound_id": current_user["compound_id"]}, {"_id": 0, "password_hash": 0})
        if not resident:
            raise HTTPException(status_code=404, detail="Resident not found")

        family = None
        family_members_list = []
        if resident.get("family_id"):
            family = await db.families.find_one({"id": resident["family_id"]}, {"_id": 0})
            if family:
                members = await db.users.find({"id": {"$in": family.get("members", [])}}, {"_id": 0, "password_hash": 0}).to_list(50)
                family_members_list = serialize_datetime(members)

        extra_members = await db.family_members.find(
            {"$or": [{"primary_resident_id": resident_id}, {"unit_id": resident_id}]}, {"_id": 0}
        ).to_list(50)

        sort_dir = -1 if sort_order == "desc" else 1
        maintenance = await db.maintenance_requests.find(
            {"$or": [{"requester_id": resident_id}, {"unit_number": resident.get("unit_number")}]}, {"_id": 0}
        ).sort("created_at", sort_dir).to_list(100)

        bookings = await db.service_bookings.find({"resident_id": resident_id}, {"_id": 0}).sort("created_at", sort_dir).to_list(100)

        invoices = []
        if resident.get("family_id"):
            invoices = await db.invoices.find(
                {"$or": [{"family_id": resident.get("family_id")}, {"resident_id": resident_id}]}, {"_id": 0}
            ).sort("created_at", sort_dir).to_list(100)

        visitors = await db.visit_requests.find(
            {"$or": [{"host_unit": resident.get("unit_number")}, {"unit_number": resident.get("unit_number")}, {"requester_id": resident_id}]}, {"_id": 0}
        ).sort("created_at", sort_dir).to_list(100)

        gate_access = await db.gate_access.find(
            {"$or": [{"unit_id": resident_id}, {"compound_id": current_user["compound_id"], "unit_number": resident.get("unit_number")}]}, {"_id": 0}
        ).sort("accessed_at", sort_dir).to_list(50)

        activities = await db.activity_logs.find(
            {"$or": [{"user_id": resident_id}, {"target_user_id": resident_id}]}, {"_id": 0}
        ).sort("timestamp", sort_dir).to_list(50)

        notifications = await db.notifications.find(
            {"$or": [{"recipient_ids": resident_id}, {"sender_id": resident_id}]}, {"_id": 0}
        ).sort("created_at", sort_dir).to_list(50)

        return serialize_datetime({
            "resident": resident,
            "family": family,
            "family_members": family_members_list,
            "extra_family_members": serialize_datetime(extra_members),
            "maintenance_requests": maintenance,
            "service_bookings": bookings,
            "invoices": invoices,
            "visitors": visitors,
            "gate_access": serialize_datetime(gate_access),
            "activities": activities,
            "notifications": notifications,
            "summary": {
                "total_family_members": len(family_members_list) + len(extra_members),
                "total_maintenance": len(maintenance),
                "open_maintenance": len([m for m in maintenance if m.get("status") in ["pending", "in_progress"]]),
                "total_bookings": len(bookings),
                "total_invoices": len(invoices),
                "pending_invoices": len([i for i in invoices if i.get("status") == "pending"]),
                "total_visitors": len(visitors),
                "total_activities": len(activities)
            }
        })
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error getting resident profile: {e}")
        raise HTTPException(status_code=500, detail="Failed to load resident profile")


@router.get("/residents/{resident_id}/export-pdf")
async def export_resident_pdf(resident_id: str, current_user: dict = Depends(require_admin)):
    await gate_company_feature(current_user, "pdf_excel_exports", "تصدير Excel/PDF")
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER

    db = get_db()
    try:
        resident = await db.users.find_one({"id": resident_id, "compound_id": current_user["compound_id"]}, {"_id": 0, "password_hash": 0})
        if not resident:
            raise HTTPException(status_code=404, detail="Resident not found")
        compound = await db.compounds.find_one({"id": current_user["compound_id"]}, {"_id": 0})

        family_members = []
        if resident.get("family_id"):
            family = await db.families.find_one({"id": resident["family_id"]}, {"_id": 0})
            if family:
                family_members = await db.users.find({"id": {"$in": family.get("members", [])}}, {"_id": 0, "password_hash": 0}).to_list(50)
        extra_members = await db.family_members.find({"$or": [{"primary_resident_id": resident_id}, {"unit_id": resident_id}]}, {"_id": 0}).to_list(50)

        maintenance = await db.maintenance_requests.find({"$or": [{"requester_id": resident_id}, {"unit_number": resident.get("unit_number")}]}, {"_id": 0}).sort("created_at", -1).to_list(50)
        bookings = await db.service_bookings.find({"resident_id": resident_id}, {"_id": 0}).sort("created_at", -1).to_list(50)
        invoices = []
        if resident.get("family_id"):
            invoices = await db.invoices.find({"$or": [{"family_id": resident.get("family_id")}, {"resident_id": resident_id}]}, {"_id": 0}).sort("created_at", -1).to_list(50)
        visitors = await db.visit_requests.find({"$or": [{"host_unit": resident.get("unit_number")}, {"unit_number": resident.get("unit_number")}]}, {"_id": 0}).sort("created_at", -1).to_list(50)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title_AR', fontName='Helvetica-Bold', fontSize=16, alignment=TA_CENTER, spaceAfter=12)
        section_style = ParagraphStyle('Section_AR', fontName='Helvetica-Bold', fontSize=13, textColor=colors.HexColor('#1e3a5f'), spaceAfter=8, spaceBefore=16)

        elements = []
        compound_name = compound.get("name", "HomeMe") if compound else "HomeMe"
        elements.append(Paragraph(f"<b>{compound_name}</b>", title_style))
        elements.append(Paragraph("Resident Profile Report", ParagraphStyle('Sub', fontName='Helvetica', fontSize=11, alignment=TA_CENTER, spaceAfter=6)))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ParagraphStyle('Date', fontName='Helvetica', fontSize=9, alignment=TA_CENTER, textColor=colors.grey, spaceAfter=20)))
        elements.append(Spacer(1, 10))

        # Personal Info
        elements.append(Paragraph("Personal Information", section_style))
        info_data = [["Field", "Value"], ["Name", resident.get("full_name", "N/A")], ["Username", resident.get("username", "N/A")], ["Email", resident.get("email", "N/A")], ["Phone", resident.get("phone", "N/A")], ["Unit", resident.get("unit_number", "N/A")], ["Role", resident.get("role", "N/A")]]
        t = Table(info_data, colWidths=[200, 300])
        t.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 9), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('PADDING', (0, 0), (-1, -1), 6)]))
        elements.append(t)
        elements.append(Spacer(1, 15))

        all_members = family_members + extra_members
        if all_members:
            elements.append(Paragraph(f"Family Members ({len(all_members)})", section_style))
            fm_data = [["Name", "Relationship", "Phone", "Email"]]
            for m in all_members:
                fm_data.append([m.get("full_name", "N/A"), m.get("relationship", m.get("role", "N/A")), m.get("phone", "N/A"), m.get("email", "N/A")])
            ft = Table(fm_data, colWidths=[140, 120, 120, 140])
            ft.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('PADDING', (0, 0), (-1, -1), 5)]))
            elements.append(ft)
            elements.append(Spacer(1, 15))

        if maintenance:
            elements.append(Paragraph(f"Maintenance ({len(maintenance)})", section_style))
            mt_data = [["Title", "Category", "Priority", "Status", "Date"]]
            for m in maintenance[:20]:
                created = m.get("created_at", "")
                if hasattr(created, 'strftime'):
                    created = created.strftime('%Y-%m-%d')
                elif isinstance(created, str):
                    created = created[:10]
                mt_data.append([str(m.get("title", "N/A"))[:30], m.get("category", "N/A"), m.get("priority", "N/A"), m.get("status", "N/A"), str(created)])
            mt = Table(mt_data, colWidths=[130, 90, 80, 80, 90])
            mt.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d97706')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('PADDING', (0, 0), (-1, -1), 5)]))
            elements.append(mt)
            elements.append(Spacer(1, 15))

        if invoices:
            elements.append(Paragraph(f"Financial ({len(invoices)})", section_style))
            inv_data = [["Description", "Amount", "Status", "Due Date"]]
            for inv in invoices[:20]:
                due = inv.get("due_date", inv.get("created_at", ""))
                if hasattr(due, 'strftime'):
                    due = due.strftime('%Y-%m-%d')
                elif isinstance(due, str):
                    due = due[:10]
                inv_data.append([str(inv.get("description", "N/A"))[:35], str(inv.get("amount", "0")), inv.get("status", "N/A"), str(due)])
            it = Table(inv_data, colWidths=[180, 90, 90, 120])
            it.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('PADDING', (0, 0), (-1, -1), 5)]))
            elements.append(it)
            elements.append(Spacer(1, 15))

        if visitors:
            elements.append(Paragraph(f"Visitors ({len(visitors)})", section_style))
            vis_data = [["Visitor", "Purpose", "Date", "Status"]]
            for v in visitors[:20]:
                vdate = v.get("visit_date", v.get("created_at", ""))
                if hasattr(vdate, 'strftime'):
                    vdate = vdate.strftime('%Y-%m-%d')
                elif isinstance(vdate, str):
                    vdate = vdate[:10]
                vis_data.append([v.get("visitor_name", "N/A"), v.get("purpose", v.get("visit_purpose", "N/A")), str(vdate), v.get("status", "N/A")])
            vt = Table(vis_data, colWidths=[140, 130, 110, 100])
            vt.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('PADDING', (0, 0), (-1, -1), 5)]))
            elements.append(vt)

        elements.append(Spacer(1, 30))
        elements.append(Paragraph(f"HomeMe - {compound_name} | Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            ParagraphStyle('Footer', fontName='Helvetica', fontSize=8, alignment=TA_CENTER, textColor=colors.grey)))

        doc.build(elements)
        pdf_content = buffer.getvalue()
        buffer.close()

        filename = f"resident_{resident.get('full_name', 'report').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return Response(content=pdf_content, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error generating resident PDF: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate PDF report")


@router.get("/reports/financial")
async def generate_financial_report(compound_id: str, start_date: str, end_date: str, language: str = "en", current_user: dict = Depends(get_current_user)):
    db = get_db()
    if current_user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    try:
        start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    except Exception:
        start = datetime.now(timezone.utc) - timedelta(days=30)
        end = datetime.now(timezone.utc)
    pdf_service = PDFReportService(db)
    pdf_content = await pdf_service.generate_financial_report(compound_id, start, end, language)
    filename = f"financial_report_{start_date[:10]}_{end_date[:10]}.pdf"
    return Response(content=pdf_content, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/reports/residents")
async def generate_residents_report(compound_id: str, language: str = "en", current_user: dict = Depends(get_current_user)):
    db = get_db()
    if current_user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    pdf_service = PDFReportService(db)
    pdf_content = await pdf_service.generate_residents_report(compound_id, language)
    filename = f"residents_report_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(content=pdf_content, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/reports/maintenance")
async def generate_maintenance_report(compound_id: str, start_date: str, end_date: str, language: str = "en", current_user: dict = Depends(get_current_user)):
    db = get_db()
    if current_user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    try:
        start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    except Exception:
        start = datetime.now(timezone.utc) - timedelta(days=30)
        end = datetime.now(timezone.utc)
    pdf_service = PDFReportService(db)
    pdf_content = await pdf_service.generate_maintenance_report(compound_id, start, end, language)
    filename = f"maintenance_report_{start_date[:10]}_{end_date[:10]}.pdf"
    return Response(content=pdf_content, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


# ══════════════════════════════════════════════════════════════════════════════
# RESIDENTS EXPORT — تصدير السكان
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/residents/export-excel")
async def export_residents_excel(
    current_user: dict = Depends(require_admin),
):
    """تصدير بيانات السكان كاملة — Excel متعدد الأوراق."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    db = get_db()
    compound_id = current_user.get("compound_id")
    if not compound_id:
        raise HTTPException(status_code=400, detail="compound_id مطلوب")

    users = await db.users.find(
        {"compound_id": compound_id},
        {"_id": 0, "password_hash": 0}
    ).to_list(5000)

    wb = openpyxl.Workbook()
    h_font = Font(bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")

    def hdr(ws, row, headers):
        ws.append(headers)
        for i in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=i)
            c.font = h_font
            c.fill = h_fill
            c.alignment = Alignment(horizontal='center')

    # Sheet 1: All Users
    ws1 = wb.active
    ws1.title = "السكان"
    ws1.sheet_view.rightToLeft = True
    hdr(ws1, 1, ["الاسم الكامل", "اسم المستخدم", "الدور", "البريد الإلكتروني", "الهاتف",
                  "رقم الوحدة", "الحالة", "تاريخ الإنشاء"])
    ROLE_AR = {"resident": "ساكن", "family_head": "رب أسرة", "admin": "مدير",
               "manager": "مشرف", "security": "أمن", "accountant": "محاسب"}
    for u in users:
        ws1.append([
            u.get("full_name", ""),
            u.get("username", ""),
            ROLE_AR.get(u.get("role", ""), u.get("role", "")),
            u.get("email", ""),
            u.get("phone", ""),
            u.get("unit_number", ""),
            "نشط" if u.get("is_active", True) else "غير نشط",
            str(u.get("created_at", ""))[:10],
        ])
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws1.column_dimensions[col].width = 22

    # Sheet 2: Residents only
    ws2 = wb.create_sheet("السكان فقط")
    ws2.sheet_view.rightToLeft = True
    hdr(ws2, 1, ["الاسم", "الهاتف", "البريد", "الوحدة", "عدد الأسرة"])
    residents = [u for u in users if u.get("role") in ("resident", "family_head")]
    for u in residents:
        family = await db.users.count_documents({"family_id": u.get("id", "")})
        ws2.append([u.get("full_name", ""), u.get("phone", ""),
                    u.get("email", ""), u.get("unit_number", ""), family])
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws2.column_dimensions[col].width = 25

    # Sheet 3: Staff
    ws3 = wb.create_sheet("الموظفون")
    ws3.sheet_view.rightToLeft = True
    hdr(ws3, 1, ["الاسم", "الدور", "الهاتف", "البريد", "تاريخ الانضمام"])
    staff = [u for u in users if u.get("role") in ("admin", "manager", "security", "accountant")]
    for u in staff:
        ws3.append([u.get("full_name", ""), ROLE_AR.get(u.get("role", ""), ""),
                    u.get("phone", ""), u.get("email", ""), str(u.get("created_at", ""))[:10]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=residents.xlsx"}
    )


@router.get("/financial/full-export-excel")
async def export_full_financial_excel(
    current_user: dict = Depends(require_admin),
):
    """تصدير شامل: السكان + الأقساط + الإيرادات + المصروفات — في ملف واحد."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    db = get_db()
    compound_id = current_user.get("compound_id")
    if not compound_id:
        raise HTTPException(status_code=400, detail="compound_id مطلوب")

    wb = openpyxl.Workbook()
    h_font = Font(bold=True, color="FFFFFF", size=11)

    def make_sheet(wb, title, color, headers, rows):
        ws = wb.create_sheet(title) if wb.worksheets else wb.active
        if not wb.worksheets or wb.active.title == "Sheet":
            ws = wb.active
            ws.title = title
        else:
            ws = wb.create_sheet(title)
        ws.sheet_view.rightToLeft = True
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.append(headers)
        for i, h in enumerate(headers, 1):
            c = ws.cell(1, i)
            c.font = h_font
            c.fill = fill
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[chr(64+i)].width = 22
        for row in rows:
            ws.append(row)
        return ws

    # 1. Installment Plans
    plans = await db.installment_plans.find({"compound_id": compound_id}, {"_id": 0}).to_list(500)
    plan_rows = []
    for p in plans:
        paid = sum(i.get("paid_amount", 0) for i in p.get("installments", []) if i.get("status") == "paid")
        overdue = sum(1 for i in p.get("installments", []) if i.get("status") != "paid" and i.get("due_date", "") < datetime.now(timezone.utc).isoformat())
        plan_rows.append([p.get("title", ""), p.get("unit_number", ""), p.get("total_amount", 0),
                          paid, p.get("total_amount", 0) - paid, p.get("deposit_amount", 0),
                          p.get("late_fee_rate", 0), overdue, p.get("status", "")])

    ws1 = wb.active
    ws1.title = "الأقساط"
    ws1.sheet_view.rightToLeft = True
    h_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    headers1 = ["الخطة", "الوحدة", "الإجمالي", "المحصّل", "المتبقي", "الوديعة", "فائدة التأخير%", "متأخر", "الحالة"]
    ws1.append(headers1)
    for i, h in enumerate(headers1, 1):
        c = ws1.cell(1, i)
        c.font = h_font; c.fill = h_fill; c.alignment = Alignment(horizontal='center')
        ws1.column_dimensions[chr(64+i)].width = 20
    for row in plan_rows:
        ws1.append(row)

    # 2. Revenue
    revenues = await db.revenue.find({"compound_id": compound_id}, {"_id": 0}).to_list(500)
    ws2 = wb.create_sheet("الإيرادات")
    ws2.sheet_view.rightToLeft = True
    h2_fill = PatternFill(start_color="1d4ed8", end_color="1d4ed8", fill_type="solid")
    headers2 = ["الوصف", "التصنيف", "المبلغ", "التاريخ", "المصدر"]
    ws2.append(headers2)
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(1, i)
        c.font = h_font; c.fill = h2_fill; c.alignment = Alignment(horizontal='center')
        ws2.column_dimensions[chr(64+i)].width = 22
    for r in revenues:
        ws2.append([r.get("description", ""), r.get("category", ""), float(r.get("amount", 0)),
                    str(r.get("date", ""))[:10], r.get("source", "")])

    # 3. Expenses
    expenses = await db.expenses.find({"compound_id": compound_id}, {"_id": 0}).to_list(500)
    ws3 = wb.create_sheet("المصروفات")
    ws3.sheet_view.rightToLeft = True
    h3_fill = PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid")
    headers3 = ["الوصف", "التصنيف", "المبلغ", "التاريخ", "الحالة"]
    ws3.append(headers3)
    for i, h in enumerate(headers3, 1):
        c = ws3.cell(1, i)
        c.font = h_font; c.fill = h3_fill; c.alignment = Alignment(horizontal='center')
        ws3.column_dimensions[chr(64+i)].width = 22
    for e in expenses:
        ws3.append([e.get("description", ""), e.get("category", ""), float(e.get("amount", 0)),
                    str(e.get("date", ""))[:10], e.get("status", "")])

    # 4. Summary
    ws4 = wb.create_sheet("الملخص")
    ws4.sheet_view.rightToLeft = True
    total_rev = sum(float(r.get("amount", 0)) for r in revenues)
    total_exp = sum(float(e.get("amount", 0)) for e in expenses)
    total_inst = sum(p.get("total_amount", 0) for p in plans)
    total_collected = sum(
        i.get("paid_amount", 0) for p in plans
        for i in p.get("installments", []) if i.get("status") == "paid"
    )
    ws4.append(["البند", "المبلغ"])
    ws4.cell(1,1).font = Font(bold=True, size=13)
    ws4.cell(1,2).font = Font(bold=True, size=13)
    for row in [
        ["إجمالي الإيرادات", total_rev],
        ["إجمالي المصروفات", total_exp],
        ["صافي الرصيد", total_rev - total_exp],
        ["", ""],
        ["إجمالي الأقساط", total_inst],
        ["المحصّل من الأقساط", total_collected],
        ["المتبقي من الأقساط", total_inst - total_collected],
    ]:
        ws4.append(row)
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=full_financial_report.xlsx"}
    )

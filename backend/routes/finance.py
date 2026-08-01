"""
Financial Management routes - Expenses, Revenue, Obligations, Balance Sheet
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from pydantic import BaseModel
from datetime import datetime, timezone
from typing import Optional, Dict
import uuid
import logging

from database import get_db
from auth_deps import get_current_user, require_admin
from helpers import serialize_datetime
from activity_logger import ActivityLogger
from financial_models import ExpenseCreate, RevenueCreate

router = APIRouter(prefix="/api")


class ObligationCreate(BaseModel):
    title: str
    description: str = ""
    total_amount: float
    month: int
    year: int
    category: str = "maintenance"
    distribution_method: str = "equal"
    unit_area_field: str = ""
    custom_amounts: Optional[Dict[str, float]] = None
    percentage_rates: Optional[Dict[str, float]] = None


@router.post("/financial/expenses")
async def create_expense(expense_data: ExpenseCreate, current_user: dict = Depends(require_admin)):
    db = get_db()
    try:
        expense = {
            "id": str(uuid.uuid4()),
            "category": expense_data.category,
            "amount": expense_data.amount,
            "description": expense_data.description,
            "date": expense_data.date,
            "payment_method": expense_data.payment_method,
            "vendor": expense_data.vendor,
            "receipt_url": expense_data.receipt_url,
            "compound_id": expense_data.compound_id,
            "created_by": current_user.get("username", ""),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "status": "completed"
        }
        await db.expenses.insert_one(expense)
        await ActivityLogger.log_activity(action_type="expense_created", username=current_user.get("username", ""), details=f"Created expense: {expense_data.description} - ${expense_data.amount}", status="success")
        expense.pop("_id", None)
        return {"success": True, "expense": expense}
    except Exception as e:
        logging.error(f"Error creating expense: {e}")
        raise HTTPException(status_code=500, detail="Failed to create expense")


@router.get("/financial/expenses")
async def get_expenses(compound_id: Optional[str] = None, start_date: Optional[str] = None, end_date: Optional[str] = None, category: Optional[str] = None, current_user: dict = Depends(require_admin)):
    db = get_db()
    try:
        query = {}
        role = current_user.get("role")
        cu_compound = current_user.get("compound_id")
        # Resolve scope: explicit query param > active compound > company-wide aggregation
        if compound_id:
            query["compound_id"] = compound_id
        elif role in ("app_owner", "super_admin"):
            pass  # global
        elif role in ("company_admin", "assistant_manager", "accountant") and current_user.get("company_id") and (not cu_compound or cu_compound in ("default-compound", "")):
            owned = await db.compounds.find(
                {"$or": [{"company_id": current_user["company_id"]}, {"management_company_id": current_user["company_id"]}]},
                {"_id": 0, "id": 1}
            ).to_list(500)
            cids = [c["id"] for c in owned if c.get("id")]
            query["compound_id"] = {"$in": cids} if cids else "__never__"
        elif cu_compound:
            query["compound_id"] = cu_compound
        if category:
            query["category"] = category
        if start_date and end_date:
            query["date"] = {"$gte": start_date, "$lte": end_date}
        expenses = await db.expenses.find(query, {"_id": 0}).sort("date", -1).to_list(length=10000)
        return {"expenses": serialize_datetime(expenses)}
    except Exception as e:
        logging.error(f"Error fetching expenses: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch expenses")


@router.post("/financial/revenue")
async def create_revenue(revenue_data: RevenueCreate, current_user: dict = Depends(require_admin)):
    db = get_db()
    try:
        revenue = {
            "id": str(uuid.uuid4()),
            "source": revenue_data.source,
            "amount": revenue_data.amount,
            "description": revenue_data.description,
            "date": revenue_data.date,
            "payment_method": revenue_data.payment_method,
            "resident_id": revenue_data.resident_id,
            "compound_id": revenue_data.compound_id,
            "created_by": current_user.get("username", ""),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "status": "completed"
        }
        await db.revenue.insert_one(revenue)
        if revenue_data.resident_id:
            await db.resident_payments.insert_one({
                "id": str(uuid.uuid4()),
                "resident_id": revenue_data.resident_id,
                "compound_id": revenue_data.compound_id,
                "amount": revenue_data.amount,
                "payment_method": revenue_data.payment_method,
                "payment_date": revenue_data.date,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
        await ActivityLogger.log_activity(action_type="revenue_created", username=current_user.get("username", ""), details=f"Created revenue: {revenue_data.description} - ${revenue_data.amount}", status="success")
        return {"success": True, "revenue": revenue}
    except Exception as e:
        logging.error(f"Error creating revenue: {e}")
        raise HTTPException(status_code=500, detail="Failed to create revenue")


@router.get("/financial/revenue")
async def get_revenue(compound_id: Optional[str] = None, start_date: Optional[str] = None, end_date: Optional[str] = None, source: Optional[str] = None, current_user: dict = Depends(require_admin)):
    db = get_db()
    try:
        query = {}
        role = current_user.get("role")
        cu_compound = current_user.get("compound_id")
        if compound_id:
            query["compound_id"] = compound_id
        elif role in ("app_owner", "super_admin"):
            pass
        elif role in ("company_admin", "assistant_manager", "accountant") and current_user.get("company_id") and (not cu_compound or cu_compound in ("default-compound", "")):
            owned = await db.compounds.find(
                {"$or": [{"company_id": current_user["company_id"]}, {"management_company_id": current_user["company_id"]}]},
                {"_id": 0, "id": 1}
            ).to_list(500)
            cids = [c["id"] for c in owned if c.get("id")]
            query["compound_id"] = {"$in": cids} if cids else "__never__"
        elif cu_compound:
            query["compound_id"] = cu_compound
        if source:
            query["source"] = source
        if start_date and end_date:
            query["date"] = {"$gte": start_date, "$lte": end_date}
        revenue = await db.revenue.find(query, {"_id": 0}).sort("date", -1).to_list(length=10000)
        return {"revenue": serialize_datetime(revenue)}
    except Exception as e:
        logging.error(f"Error fetching revenue: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch revenue")


@router.get("/financial/reports/summary")
async def get_financial_summary(compound_id: str, start_date: str, end_date: str, current_user: dict = Depends(require_admin)):
    db = get_db()
    try:
        expenses = await db.expenses.find({"compound_id": compound_id, "date": {"$gte": start_date, "$lte": end_date}}).to_list(length=10000)
        revenue = await db.revenue.find({"compound_id": compound_id, "date": {"$gte": start_date, "$lte": end_date}}).to_list(length=10000)
        total_expenses = sum(e["amount"] for e in expenses)
        total_revenue = sum(r["amount"] for r in revenue)
        net_profit = total_revenue - total_expenses
        profit_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0
        expenses_by_category = {}
        for e in expenses:
            cat = e["category"]
            expenses_by_category[cat] = expenses_by_category.get(cat, 0) + e["amount"]
        revenue_by_source = {}
        for r in revenue:
            src = r["source"]
            revenue_by_source[src] = revenue_by_source.get(src, 0) + r["amount"]
        return {"period": "custom", "start_date": start_date, "end_date": end_date, "total_expenses": total_expenses, "total_revenue": total_revenue, "net_profit": net_profit, "profit_margin": round(profit_margin, 2), "expenses_by_category": expenses_by_category, "revenue_by_source": revenue_by_source}
    except Exception as e:
        logging.error(f"Error generating financial summary: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate summary")


@router.post("/financial/obligations")
async def create_obligation(data: ObligationCreate, current_user: dict = Depends(require_admin)):
    db = get_db()
    try:
        compound_id = current_user["compound_id"]
        families = await db.families.find({"compound_id": compound_id}, {"_id": 0}).to_list(100)
        if not families:
            raise HTTPException(status_code=400, detail="No units found in compound")

        unit_count = len(families)
        family_heads = {}
        for family in families:
            head = await db.users.find_one({"id": {"$in": family.get("members", [])}, "is_family_head": True}, {"_id": 0, "id": 1, "full_name": 1, "unit_number": 1, "unit_area": 1})
            if not head:
                head = await db.users.find_one({"id": {"$in": family.get("members", [])}}, {"_id": 0, "id": 1, "full_name": 1, "unit_number": 1, "unit_area": 1})
            if head:
                family_heads[family["id"]] = head

        unit_amounts = {}
        method_label = "بالتساوي"

        if data.distribution_method == "equal":
            per_unit = round(data.total_amount / unit_count, 2)
            for fid in family_heads:
                unit_amounts[fid] = per_unit
            method_label = "بالتساوي"
        elif data.distribution_method == "per_sqm":
            total_area = sum(float(head.get("unit_area", 100)) for head in family_heads.values())
            if total_area > 0:
                for fid, head in family_heads.items():
                    area = float(head.get("unit_area", 100))
                    unit_amounts[fid] = round((area / total_area) * data.total_amount, 2)
            else:
                per_unit = round(data.total_amount / unit_count, 2)
                for fid in family_heads:
                    unit_amounts[fid] = per_unit
            method_label = "حسب المساحة"
        elif data.distribution_method == "percentage":
            if data.percentage_rates:
                for fid, pct in data.percentage_rates.items():
                    if fid in family_heads:
                        unit_amounts[fid] = round(data.total_amount * (pct / 100), 2)
            assigned = sum(unit_amounts.values())
            remaining = [fid for fid in family_heads if fid not in unit_amounts]
            if remaining and assigned < data.total_amount:
                each = round((data.total_amount - assigned) / len(remaining), 2)
                for fid in remaining:
                    unit_amounts[fid] = each
            method_label = "نسبة مئوية"
        elif data.distribution_method == "custom":
            if data.custom_amounts:
                for fid, amt in data.custom_amounts.items():
                    if fid in family_heads:
                        unit_amounts[fid] = float(amt)
            assigned = sum(unit_amounts.values())
            remaining = [fid for fid in family_heads if fid not in unit_amounts]
            if remaining:
                each = round(max(0, data.total_amount - assigned) / len(remaining), 2)
                for fid in remaining:
                    unit_amounts[fid] = each
            method_label = "مبلغ مخصص"
        else:
            per_unit = round(data.total_amount / unit_count, 2)
            for fid in family_heads:
                unit_amounts[fid] = per_unit

        obligation_id = str(uuid.uuid4())
        obligation = {
            "id": obligation_id, "compound_id": compound_id, "title": data.title, "description": data.description,
            "total_amount": data.total_amount, "distribution_method": data.distribution_method,
            "distribution_label": method_label, "unit_count": unit_count,
            "month": data.month, "year": data.year, "category": data.category,
            "created_by": current_user["id"], "created_at": datetime.now(timezone.utc)
        }
        await db.obligations.insert_one(obligation)

        charges_created = 0
        for family in families:
            fid = family["id"]
            head = family_heads.get(fid)
            if not head:
                continue
            amount = unit_amounts.get(fid, 0)
            if amount <= 0:
                continue
            charge = {
                "id": str(uuid.uuid4()), "obligation_id": obligation_id, "compound_id": compound_id,
                "resident_id": head["id"], "resident_name": head.get("full_name", ""),
                "unit_number": head.get("unit_number", ""), "family_id": fid,
                "title": data.title, "category": data.category, "amount": amount,
                "month": data.month, "year": data.year, "status": "pending",
                "paid_at": None, "created_at": datetime.now(timezone.utc)
            }
            await db.unit_charges.insert_one(charge)
            charges_created += 1

        await db.expenses.insert_one({
            "id": str(uuid.uuid4()), "category": data.category,
            "amount": data.total_amount, "description": data.title + " - " + data.description,
            "date": datetime.now(timezone.utc).isoformat(), "payment_method": "other",
            "compound_id": compound_id, "obligation_id": obligation_id,
            "created_by": current_user.get("username", ""),
            "created_at": datetime.now(timezone.utc).isoformat(), "status": "completed"
        })

        return {
            "message": f"تم إنشاء الالتزام وتوزيعه على {charges_created} وحدة ({method_label})",
            "obligation_id": obligation_id, "distribution_method": data.distribution_method,
            "units_charged": charges_created,
            "unit_amounts": {family_heads[fid].get("unit_number", fid): amt for fid, amt in unit_amounts.items() if fid in family_heads}
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating obligation: {e}")
        raise HTTPException(status_code=500, detail="Failed to create obligation")


@router.get("/financial/obligations")
async def get_obligations(month: Optional[int] = None, year: Optional[int] = None, current_user: dict = Depends(require_admin)):
    db = get_db()
    try:
        query = {"compound_id": current_user["compound_id"]}
        if month:
            query["month"] = month
        if year:
            query["year"] = year
        obligations = await db.obligations.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
        return {"obligations": serialize_datetime(obligations)}
    except Exception as e:
        logging.error(f"Error fetching obligations: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch obligations")


@router.get("/financial/unit-charges")
async def get_unit_charges(obligation_id: Optional[str] = None, month: Optional[int] = None, year: Optional[int] = None, status: Optional[str] = None, current_user: dict = Depends(require_admin)):
    db = get_db()
    try:
        query = {"compound_id": current_user["compound_id"]}
        if obligation_id:
            query["obligation_id"] = obligation_id
        if month:
            query["month"] = month
        if year:
            query["year"] = year
        if status:
            query["status"] = status
        charges = await db.unit_charges.find(query, {"_id": 0}).sort("unit_number", 1).to_list(500)
        total = len(charges)
        paid = len([c for c in charges if c.get("status") == "paid"])
        unpaid = total - paid
        total_amount = sum(c.get("amount", 0) for c in charges)
        paid_amount = sum(c.get("amount", 0) for c in charges if c.get("status") == "paid")
        return {
            "charges": serialize_datetime(charges),
            "summary": {"total": total, "paid": paid, "unpaid": unpaid, "total_amount": total_amount, "paid_amount": paid_amount, "unpaid_amount": total_amount - paid_amount}
        }
    except Exception as e:
        logging.error(f"Error fetching unit charges: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch unit charges")


@router.put("/financial/unit-charges/{charge_id}/pay")
async def mark_charge_paid(charge_id: str, current_user: dict = Depends(require_admin)):
    db = get_db()
    try:
        charge = await db.unit_charges.find_one({"id": charge_id, "compound_id": current_user["compound_id"]})
        if not charge:
            raise HTTPException(status_code=404, detail="Charge not found")
        now = datetime.now(timezone.utc)
        await db.unit_charges.update_one({"id": charge_id}, {"$set": {"status": "paid", "paid_at": now, "paid_by": current_user["id"]}})
        await db.revenue.insert_one({
            "id": str(uuid.uuid4()), "source": "maintenance_fees",
            "amount": charge["amount"],
            "description": f"سداد {charge.get('title', '')} - وحدة {charge.get('unit_number', '')}",
            "date": now.isoformat(), "payment_method": "other",
            "resident_id": charge.get("resident_id"), "compound_id": current_user["compound_id"],
            "charge_id": charge_id, "created_by": current_user.get("username", ""),
            "created_at": now.isoformat(), "status": "completed"
        })
        return {"message": "تم تسجيل السداد بنجاح", "charge_id": charge_id}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error marking charge paid: {e}")
        raise HTTPException(status_code=500, detail="Failed to update charge")


@router.post("/financial/unit-charges/notify-unpaid")
async def notify_unpaid_units(obligation_id: str = "", month: Optional[int] = None, year: Optional[int] = None, current_user: dict = Depends(require_admin)):
    db = get_db()
    try:
        query = {"compound_id": current_user["compound_id"], "status": "pending"}
        if obligation_id:
            query["obligation_id"] = obligation_id
        if month:
            query["month"] = month
        if year:
            query["year"] = year
        unpaid = await db.unit_charges.find(query, {"_id": 0}).to_list(500)
        notified = 0
        for charge in unpaid:
            resident_id = charge.get("resident_id")
            if not resident_id:
                continue
            notification = {
                "id": str(uuid.uuid4()), "compound_id": current_user["compound_id"],
                "sender_id": "system", "title": "تذكير بسداد التزام",
                "content": f"يرجى سداد التزام '{charge.get('title', '')}' بمبلغ {charge.get('amount', 0)} - شهر {charge.get('month')}/{charge.get('year')}",
                "type": "payment_reminder", "recipient_ids": [resident_id],
                "is_read": False, "created_at": datetime.now(timezone.utc)
            }
            await db.notifications.insert_one(notification)
            notified += 1
        return {"message": f"تم إرسال {notified} إشعار للوحدات المتأخرة", "notified_count": notified}
    except Exception as e:
        logging.error(f"Error notifying unpaid: {e}")
        raise HTTPException(status_code=500, detail="Failed to send notifications")


@router.get("/financial/balance-sheet")
async def get_balance_sheet(year: Optional[int] = None, compound_id: Optional[str] = None, current_user: dict = Depends(require_admin)):
    db = get_db()
    try:
        role = current_user.get("role")
        cu_compound = current_user.get("compound_id")
        # Resolve scope: explicit query → active compound (X-Active-Compound-Id) → company-wide aggregation
        if compound_id:
            scope = {"compound_id": compound_id}
            label_compound = compound_id
        elif role in ("company_admin", "assistant_manager", "accountant") and current_user.get("company_id") and (not cu_compound or cu_compound in ("default-compound", "")):
            # Aggregate across ALL company's compounds when no specific compound is active
            company_id = current_user["company_id"]
            owned = await db.compounds.find(
                {"$or": [{"company_id": company_id}, {"management_company_id": company_id}]},
                {"_id": 0, "id": 1}
            ).to_list(500)
            cids = [c["id"] for c in owned if c.get("id")]
            scope = {"compound_id": {"$in": cids}} if cids else {"compound_id": "__never__"}
            label_compound = f"company:{company_id}"
        elif role in ("app_owner", "super_admin") and not cu_compound:
            scope = {}  # global
            label_compound = "all"
        else:
            scope = {"compound_id": cu_compound or "default-compound"}
            label_compound = cu_compound or "default-compound"
        current_year = year or datetime.now().year
        expenses = await db.expenses.find(scope, {"_id": 0}).to_list(2000)
        total_expenses = sum(float(e.get("amount", 0)) for e in expenses)
        exp_by_cat = {}
        for e in expenses:
            cat = e.get("category", "other")
            exp_by_cat[cat] = exp_by_cat.get(cat, 0) + float(e.get("amount", 0))
        revenues = await db.revenue.find(scope, {"_id": 0}).to_list(2000)
        total_revenue = sum(float(r.get("amount", 0)) for r in revenues)
        rev_by_src = {}
        for r in revenues:
            src = r.get("source", "other")
            rev_by_src[src] = rev_by_src.get(src, 0) + float(r.get("amount", 0))
        all_charges = await db.unit_charges.find(scope, {"_id": 0}).to_list(5000)
        total_charged = sum(float(c.get("amount", 0)) for c in all_charges)
        total_collected = sum(float(c.get("amount", 0)) for c in all_charges if c.get("status") == "paid")
        monthly = {}
        for e in expenses:
            date_str = e.get("date", e.get("created_at", ""))
            if date_str:
                m = date_str[:7]
                if m not in monthly:
                    monthly[m] = {"expenses": 0, "revenue": 0}
                monthly[m]["expenses"] += float(e.get("amount", 0))
        for r in revenues:
            date_str = r.get("date", r.get("created_at", ""))
            if date_str:
                m = date_str[:7]
                if m not in monthly:
                    monthly[m] = {"expenses": 0, "revenue": 0}
                monthly[m]["revenue"] += float(r.get("amount", 0))
        return {
            "compound_id": label_compound, "year": current_year,
            "total_expenses": round(total_expenses, 2), "total_revenue": round(total_revenue, 2),
            "net_balance": round(total_revenue - total_expenses, 2),
            "expenses_by_category": exp_by_cat, "revenue_by_source": rev_by_src,
            "obligations": {"total_charged": round(total_charged, 2), "total_collected": round(total_collected, 2), "total_outstanding": round(total_charged - total_collected, 2), "collection_rate": round((total_collected / total_charged * 100) if total_charged > 0 else 0, 1)},
            "monthly_breakdown": dict(sorted(monthly.items())),
            "recent_expenses": serialize_datetime(expenses[:10]),
            "recent_revenue": serialize_datetime(revenues[:10])
        }
    except Exception as e:
        logging.error(f"Error getting balance sheet: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate balance sheet")


@router.get("/financial/balance-sheet/export-pdf")
async def export_balance_sheet_pdf(year: Optional[int] = None, compound_id: Optional[str] = None, current_user: dict = Depends(require_admin)):
    """Export the full balance sheet (revenue + expenses + breakdowns + monthly trend) as a branded Arabic PDF."""
    from fastapi.responses import Response
    from services.pdf_report_service import render_balance_sheet

    db = get_db()
    try:
        # Reuse the same scoping logic as get_balance_sheet
        role = current_user.get("role")
        cu_compound = current_user.get("compound_id")
        compound_label = ""
        if compound_id:
            scope = {"compound_id": compound_id}
            cmpd = await db.compounds.find_one({"id": compound_id}, {"_id": 0, "name": 1})
            compound_label = (cmpd or {}).get("name") or compound_id
        elif role in ("company_admin", "assistant_manager", "accountant") and current_user.get("company_id") and (not cu_compound or cu_compound in ("default-compound", "")):
            company_id = current_user["company_id"]
            owned = await db.compounds.find(
                {"$or": [{"company_id": company_id}, {"management_company_id": company_id}]},
                {"_id": 0, "id": 1}
            ).to_list(500)
            cids = [c["id"] for c in owned if c.get("id")]
            scope = {"compound_id": {"$in": cids}} if cids else {"compound_id": "__never__"}
            company_doc = await db.companies.find_one({"id": company_id}, {"_id": 0, "name": 1}) or {}
            compound_label = f"{company_doc.get('name', 'الشركة')} (إجمالي {len(cids)} كمبوندات)"
        elif role in ("app_owner", "super_admin") and not cu_compound:
            scope = {}
            compound_label = "كل الكمبوندات"
        else:
            scope = {"compound_id": cu_compound or "default-compound"}
            cmpd = await db.compounds.find_one({"id": scope["compound_id"]}, {"_id": 0, "name": 1})
            compound_label = (cmpd or {}).get("name") or "الكمبوند"

        current_year = year or datetime.now().year
        period = f"عام {current_year}"

        expenses = await db.expenses.find(scope, {"_id": 0}).to_list(2000)
        total_expenses = sum(float(e.get("amount", 0)) for e in expenses)
        exp_by_cat: Dict[str, float] = {}
        for e in expenses:
            cat = e.get("category", "other")
            exp_by_cat[cat] = exp_by_cat.get(cat, 0) + float(e.get("amount", 0))

        revenues = await db.revenue.find(scope, {"_id": 0}).to_list(2000)
        total_revenue = sum(float(r.get("amount", 0)) for r in revenues)
        rev_by_src: Dict[str, float] = {}
        for r in revenues:
            src = r.get("source", "other")
            rev_by_src[src] = rev_by_src.get(src, 0) + float(r.get("amount", 0))

        all_charges = await db.unit_charges.find(scope, {"_id": 0}).to_list(5000)
        total_charged = sum(float(c.get("amount", 0)) for c in all_charges)
        total_collected = sum(float(c.get("amount", 0)) for c in all_charges if c.get("status") == "paid")
        coll_rate = round((total_collected / total_charged * 100) if total_charged > 0 else 0, 1)

        # Monthly breakdown
        monthly: Dict[str, Dict[str, float]] = {}
        for e in expenses:
            d = e.get("date") or e.get("created_at") or ""
            if isinstance(d, str) and d:
                m = d[:7]
                monthly.setdefault(m, {"expenses": 0, "revenue": 0})
                monthly[m]["expenses"] += float(e.get("amount", 0))
        for r in revenues:
            d = r.get("date") or r.get("created_at") or ""
            if isinstance(d, str) and d:
                m = d[:7]
                monthly.setdefault(m, {"expenses": 0, "revenue": 0})
                monthly[m]["revenue"] += float(r.get("amount", 0))

        # Sort recent
        recent_expenses = sorted(expenses, key=lambda x: x.get("date") or x.get("created_at") or "", reverse=True)
        recent_revenue = sorted(revenues, key=lambda x: x.get("date") or x.get("created_at") or "", reverse=True)

        # Branding
        branding = None
        try:
            from services.branding import get_compound_branding
            single_cid = compound_id or (cu_compound if cu_compound and cu_compound != "default-compound" else None)
            if single_cid:
                cmpd_doc = await db.compounds.find_one({"id": single_cid}, {"_id": 0})
                branding = get_compound_branding(cmpd_doc)
        except Exception:
            pass

        pdf_bytes = render_balance_sheet(
            compound_name=compound_label,
            period=period,
            total_revenue=total_revenue,
            total_expenses=total_expenses,
            expenses_by_category=exp_by_cat,
            revenue_by_source=rev_by_src,
            monthly_breakdown=monthly,
            obligations={"total_charged": total_charged, "total_collected": total_collected, "collection_rate": coll_rate},
            recent_expenses=recent_expenses,
            recent_revenue=recent_revenue,
            currency="EGP",
            branding=branding,
        )
        filename = f"balance_sheet_{current_year}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        logging.error(f"Error exporting balance sheet PDF: {e}")
        raise HTTPException(status_code=500, detail="Failed to export balance sheet PDF")


@router.get("/financial/residents/{resident_id}/account")
async def get_resident_account(resident_id: str, current_user: dict = Depends(get_current_user)):
    db = get_db()
    if current_user.get("role") != "admin" and current_user["id"] != resident_id:
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        resident = await db.users.find_one({"id": resident_id})
        if not resident:
            raise HTTPException(status_code=404, detail="Resident not found")
        charges = await db.resident_charges.find({"resident_id": resident_id}).to_list(length=10000)
        payments = await db.resident_payments.find({"resident_id": resident_id}).sort("payment_date", -1).to_list(length=10000)
        total_charges = sum(c["amount"] for c in charges)
        total_payments = sum(p["amount"] for p in payments)
        pending_charges = [c for c in charges if c.get("status") == "pending"]
        return {"resident_id": resident_id, "resident_name": resident.get("full_name"), "unit_number": resident.get("unit_number"), "total_charges": total_charges, "total_payments": total_payments, "balance": total_charges - total_payments, "pending_charges": pending_charges, "recent_payments": payments[:10]}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error fetching resident account: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch account")


@router.post("/financial/residents/payments")
async def create_resident_payment(payment_data: dict, current_user: dict = Depends(get_current_user)):
    db = get_db()
    try:
        resident_id = payment_data.get("resident_id")
        if current_user.get("role") != "admin" and current_user["id"] != resident_id:
            raise HTTPException(status_code=403, detail="Access denied")
        payment = {
            "id": str(uuid.uuid4()), "resident_id": resident_id,
            "compound_id": payment_data.get("compound_id"), "amount": payment_data.get("amount"),
            "payment_method": payment_data.get("payment_method"),
            "payment_date": datetime.now(timezone.utc).isoformat(),
            "reference": payment_data.get("reference"), "notes": payment_data.get("notes"),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.resident_payments.insert_one(payment)
        if payment_data.get("reference") and payment_data["reference"].startswith("CHARGE-"):
            charge_id = payment_data["reference"].replace("CHARGE-", "")
            await db.resident_charges.update_one({"id": charge_id}, {"$set": {"status": "paid"}})
        await ActivityLogger.log_activity(action_type="payment_created", username=current_user.get("username", ""), details=f"Payment of ${payment_data.get('amount')} made", status="success")
        return {"success": True, "payment": payment}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating payment: {e}")
        raise HTTPException(status_code=500, detail="Failed to create payment")


@router.get("/financial/residents/payments/{payment_id}/receipt")
async def get_payment_receipt(payment_id: str, current_user: dict = Depends(get_current_user)):
    db = get_db()
    try:
        payment = await db.resident_payments.find_one({"id": payment_id})
        if not payment:
            raise HTTPException(status_code=404, detail="Payment not found")
        if current_user.get("role") != "admin" and current_user["id"] != payment.get("resident_id"):
            raise HTTPException(status_code=403, detail="Access denied")
        receipt_text = f"PAYMENT RECEIPT\nReceipt ID: {payment_id}\nDate: {payment.get('payment_date')}\nAmount: ${payment.get('amount')}\nMethod: {payment.get('payment_method')}"
        return {"receipt": receipt_text, "payment": payment}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error generating receipt: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate receipt")


# ══════════════════════════════════════════════════════════════════════════════
# INSTALLMENT PLANS — أقساط الوحدات
# ══════════════════════════════════════════════════════════════════════════════

class InstallmentPlanCreate(BaseModel):
    resident_id: str
    unit_number: str
    title: str                          # "ثمن الوحدة / رسوم التسجيل / ..."
    total_amount: float
    down_payment: float = 0.0           # دفعة أولى
    installment_count: int              # عدد الأقساط
    installment_amount: float           # قيمة القسط الشهري
    start_date: str                     # "2026-08-01"
    late_fee_rate: float = 0.0          # % فائدة تأخير شهرية
    early_payment_discount: float = 0.0 # % خصم الدفع الكاش
    deposit_amount: float = 0.0         # وديعة
    deposit_refundable: bool = True
    notes: str = ""
    pricing_method: str = "fixed"       # fixed | per_sqm | percentage | custom


@router.post("/financial/installment-plans")
async def create_installment_plan(
    data: InstallmentPlanCreate,
    current_user: dict = Depends(require_admin),
):
    """إنشاء خطة أقساط لساكن/وحدة."""
    db = get_db()
    compound_id = current_user.get("compound_id")

    # Build installment schedule
    from datetime import datetime, timezone, timedelta
    import calendar

    base_date = datetime.fromisoformat(data.start_date)
    installments = []
    for i in range(data.installment_count):
        # Add months
        month = base_date.month + i
        year = base_date.year + (month - 1) // 12
        month = (month - 1) % 12 + 1
        due_date = base_date.replace(year=year, month=month)
        installments.append({
            "number": i + 1,
            "due_date": due_date.isoformat(),
            "amount": data.installment_amount,
            "status": "pending",
            "paid_at": None,
            "paid_amount": None,
            "late_fee": 0.0,
            "discount": 0.0,
        })

    plan = {
        "id": str(uuid.uuid4()),
        "compound_id": compound_id,
        "resident_id": data.resident_id,
        "unit_number": data.unit_number,
        "title": data.title,
        "total_amount": data.total_amount,
        "down_payment": data.down_payment,
        "installment_count": data.installment_count,
        "installment_amount": data.installment_amount,
        "late_fee_rate": data.late_fee_rate,
        "early_payment_discount": data.early_payment_discount,
        "deposit_amount": data.deposit_amount,
        "deposit_refundable": data.deposit_refundable,
        "deposit_status": "held" if data.deposit_amount > 0 else None,
        "pricing_method": data.pricing_method,
        "notes": data.notes,
        "status": "active",
        "installments": installments,
        "created_by": current_user.get("id"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.installment_plans.insert_one(plan)
    plan.pop("_id", None)
    return {"success": True, "plan": plan}


@router.get("/financial/installment-plans")
async def get_installment_plans(
    compound_id: str = None,
    resident_id: str = None,
    current_user: dict = Depends(get_current_user),
):
    """قائمة خطط الأقساط — يدعم الأدمن وشركات الإدارة."""
    db = get_db()
    role = current_user.get("role", "")

    # Scope by role
    if role in ("app_owner", "super_admin"):
        # Can see all or filter by compound
        cid = compound_id
        query = {"compound_id": cid} if cid else {}
    elif role == "company_admin" and current_user.get("company_id"):
        # Get all compounds for this company
        compounds = await db.compounds.find(
            {"$or": [
                {"company_id": current_user["company_id"]},
                {"management_company_id": current_user["company_id"]}
            ]}, {"_id": 0, "id": 1}
        ).to_list(200)
        compound_ids = [c["id"] for c in compounds]
        query = {"compound_id": {"$in": compound_ids}}
        if compound_id:
            query = {"compound_id": compound_id}
    else:
        cid = compound_id or current_user.get("compound_id")
        query = {"compound_id": cid}
    if resident_id:
        query["resident_id"] = resident_id
    plans = await db.installment_plans.find(query, {"_id": 0}).to_list(length=500)
    return {"plans": plans, "total": len(plans)}


@router.get("/financial/installment-plans/{plan_id}")
async def get_installment_plan(
    plan_id: str,
    current_user: dict = Depends(get_current_user),
):
    db = get_db()
    plan = await db.installment_plans.find_one({"id": plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="الخطة غير موجودة")
    return plan


@router.put("/financial/installment-plans/{plan_id}/pay/{installment_number}")
async def pay_installment(
    plan_id: str,
    installment_number: int,
    body: dict,
    current_user: dict = Depends(require_admin),
):
    """تسجيل دفع قسط مع حساب فائدة التأخير / خصم الكاش."""
    db = get_db()
    plan = await db.installment_plans.find_one({"id": plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="الخطة غير موجودة")

    from datetime import datetime, timezone, date
    payment_method = body.get("payment_method", "cash")
    paid_date = datetime.now(timezone.utc)

    installments = plan.get("installments", [])
    idx = next((i for i, x in enumerate(installments) if x["number"] == installment_number), None)
    if idx is None:
        raise HTTPException(status_code=404, detail="القسط غير موجود")

    inst = installments[idx]
    base_amount = inst["amount"]

    # Calculate late fee
    late_fee = 0.0
    if plan.get("late_fee_rate", 0) > 0:
        due = datetime.fromisoformat(inst["due_date"])
        if paid_date > due:
            months_late = max(1, (paid_date.year - due.year) * 12 + paid_date.month - due.month)
            late_fee = round(base_amount * (plan["late_fee_rate"] / 100) * months_late, 2)

    # Calculate cash discount
    discount = 0.0
    if payment_method == "cash" and plan.get("early_payment_discount", 0) > 0:
        discount = round(base_amount * (plan["early_payment_discount"] / 100), 2)

    final_amount = base_amount + late_fee - discount

    installments[idx].update({
        "status": "paid",
        "paid_at": paid_date.isoformat(),
        "paid_amount": final_amount,
        "late_fee": late_fee,
        "discount": discount,
        "payment_method": payment_method,
        "notes": body.get("notes", ""),
    })

    # Check if all paid
    all_paid = all(x["status"] == "paid" for x in installments)

    await db.installment_plans.update_one(
        {"id": plan_id},
        {"$set": {
            "installments": installments,
            "status": "completed" if all_paid else "active",
        }}
    )

    return {
        "success": True,
        "installment_number": installment_number,
        "base_amount": base_amount,
        "late_fee": late_fee,
        "discount": discount,
        "final_amount": final_amount,
        "all_paid": all_paid,
    }


@router.put("/financial/installment-plans/{plan_id}")
async def update_installment_plan(
    plan_id: str,
    body: dict,
    current_user: dict = Depends(require_admin),
):
    """تعديل خطة أقساط."""
    db = get_db()
    allowed = ["title", "notes", "late_fee_rate", "early_payment_discount",
               "deposit_amount", "deposit_status", "status"]
    update = {k: v for k, v in body.items() if k in allowed}
    if update:
        await db.installment_plans.update_one({"id": plan_id}, {"$set": update})
    return {"success": True}


@router.delete("/financial/installment-plans/{plan_id}")
async def delete_installment_plan(
    plan_id: str,
    current_user: dict = Depends(require_admin),
):
    """حذف خطة أقساط."""
    db = get_db()
    await db.installment_plans.delete_one({"id": plan_id})
    return {"success": True}


@router.get("/financial/installment-plans/export/excel")
async def export_installment_plans_excel(
    compound_id: str = None,
    current_user: dict = Depends(require_admin),
):
    """تصدير خطط الأقساط Excel."""
    import openpyxl
    from fastapi.responses import Response
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    db = get_db()
    cid = compound_id or current_user.get("compound_id")
    plans = await db.installment_plans.find({"compound_id": cid}, {"_id": 0}).to_list(500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "خطط الأقساط"

    headers = ["الساكن", "الوحدة", "العنوان", "الإجمالي", "الأقساط", "قيمة القسط",
               "فائدة التأخير%", "خصم الكاش%", "الوديعة", "الحالة", "ملاحظات"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="059669")

    for row, p in enumerate(plans, 2):
        ws.append([
            p.get("resident_id", ""), p.get("unit_number", ""), p.get("title", ""),
            p.get("total_amount", 0), p.get("installment_count", 0),
            p.get("installment_amount", 0), p.get("late_fee_rate", 0),
            p.get("early_payment_discount", 0), p.get("deposit_amount", 0),
            p.get("status", ""), p.get("notes", ""),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=installment_plans.xlsx"}
    )


@router.get("/financial/residents/{resident_id}/full-account")
async def get_resident_full_account(
    resident_id: str,
    current_user: dict = Depends(get_current_user),
):
    """الحساب المالي الكامل للساكن — كل الالتزامات والأقساط والودائع."""
    db = get_db()
    compound_id = current_user.get("compound_id")

    # Installment plans
    plans = await db.installment_plans.find(
        {"resident_id": resident_id}, {"_id": 0}
    ).to_list(100)

    # Unit charges (maintenance, shared expenses)
    charges = await db.unit_charges.find(
        {"resident_id": resident_id, "compound_id": compound_id}, {"_id": 0}
    ).to_list(200)

    # Payments
    payments = await db.payments.find(
        {"resident_id": resident_id, "compound_id": compound_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)

    # Calculate totals
    total_installments = sum(p.get("total_amount", 0) for p in plans)
    total_paid_installments = sum(
        i.get("paid_amount", 0)
        for p in plans for i in p.get("installments", [])
        if i.get("status") == "paid"
    )
    total_deposits = sum(p.get("deposit_amount", 0) for p in plans)
    total_charges = sum(c.get("amount", 0) for c in charges)
    total_paid_charges = sum(c.get("amount", 0) for c in charges if c.get("status") == "paid")
    total_late_fees = sum(
        i.get("late_fee", 0)
        for p in plans for i in p.get("installments", [])
    )
    total_discounts = sum(
        i.get("discount", 0)
        for p in plans for i in p.get("installments", [])
    )

    # Pending installments
    pending_installments = [
        {**i, "plan_title": p.get("title"), "plan_id": p.get("id"),
         "late_fee_rate": p.get("late_fee_rate", 0),
         "early_payment_discount": p.get("early_payment_discount", 0)}
        for p in plans for i in p.get("installments", [])
        if i.get("status") == "pending"
    ]

    return {
        "resident_id": resident_id,
        "installment_plans": plans,
        "charges": charges,
        "payments": payments,
        "pending_installments": sorted(pending_installments, key=lambda x: x.get("due_date", "")),
        "summary": {
            "total_installments": total_installments,
            "total_paid_installments": total_paid_installments,
            "total_remaining_installments": total_installments - total_paid_installments,
            "total_deposits": total_deposits,
            "total_charges": total_charges,
            "total_paid_charges": total_paid_charges,
            "total_remaining_charges": total_charges - total_paid_charges,
            "total_late_fees": total_late_fees,
            "total_discounts": total_discounts,
            "grand_total_due": (total_installments - total_paid_installments) + (total_charges - total_paid_charges) + total_late_fees,
        }
    }


# ══════════════════════════════════════════════════════════════════════════════
# NOTIFICATIONS — إشعارات الأقساط المتأخرة
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/financial/installment-plans/notify-overdue")
async def notify_overdue_installments(
    body: dict,
    current_user: dict = Depends(require_admin),
):
    """إرسال إشعارات للسكان المتأخرين في دفع الأقساط."""
    from datetime import datetime, timezone
    from email_service import email_service

    db = get_db()
    compound_id = body.get("compound_id") or current_user.get("compound_id")
    query = {"compound_id": compound_id} if compound_id else {}
    plans = await db.installment_plans.find(query, {"_id": 0}).to_list(500)

    now = datetime.now(timezone.utc)
    notified = 0

    for plan in plans:
        overdue = [i for i in plan.get("installments", [])
                   if i.get("status") != "paid" and
                   datetime.fromisoformat(i["due_date"].replace("Z", "+00:00")) < now]
        if not overdue:
            continue

        # Get resident email
        resident = await db.users.find_one(
            {"id": plan["resident_id"]}, {"_id": 0, "email": 1, "full_name": 1, "username": 1}
        )
        if not resident or not resident.get("email"):
            continue

        total_overdue = sum(i["amount"] for i in overdue)
        late_fees = sum(
            i["amount"] * (plan.get("late_fee_rate", 0) / 100)
            for i in overdue
        )

        html = f"""
        <div dir='rtl' style='font-family:Cairo,Arial,sans-serif;max-width:600px;margin:auto'>
          <div style='background:#dc2626;color:#fff;padding:20px;border-radius:12px 12px 0 0;text-align:center'>
            <h2 style='margin:0'>⚠️ تنبيه: أقساط متأخرة</h2>
          </div>
          <div style='background:#fff;padding:20px;border:1px solid #e5e7eb;border-radius:0 0 12px 12px'>
            <p>مرحباً <strong>{resident.get('full_name') or resident.get('username')}</strong>،</p>
            <p>يوجد لديك <strong>{len(overdue)} قسط متأخر</strong> بإجمالي <strong>{total_overdue:,.0f} ج.م</strong></p>
            <p>الخطة: <strong>{plan.get('title')}</strong> — وحدة {plan.get('unit_number', '')}</p>
            {f"<p style='color:#dc2626'>فوائد التأخير المستحقة: {late_fees:,.0f} ج.م</p>" if late_fees > 0 else ""}
            <p>يرجى سداد المبلغ في أقرب وقت ممكن لتجنب تراكم الفوائد.</p>
            <div style='background:#fef2f2;border:1px solid #fecaca;border-radius:8px;padding:12px;margin-top:16px'>
              <p style='margin:0;font-size:12px;color:#991b1b'>
                للدفع: تواصل مع إدارة الكمبوند أو ادفع عبر التطبيق
              </p>
            </div>
          </div>
        </div>"""

        try:
            await email_service.send_email(
                to_email=resident["email"],
                subject=f"⚠️ تنبيه أقساط متأخرة — {plan.get('title')}",
                html_content=html,
            )
            notified += 1
        except Exception as _e:
            pass

    return {"success": True, "notified": notified, "message": f"تم إرسال {notified} إشعار"}


@router.post("/financial/installment-plans/{plan_id}/upload-proof")
async def upload_installment_proof(
    plan_id: str,
    installment_number: int = Form(...),
    proof: UploadFile = File(...),
    notes: str = Form(""),
    current_user: dict = Depends(get_current_user),
):
    """الساكن يرفع إيصال دفع قسط."""
    import os
    db = get_db()
    plan = await db.installment_plans.find_one({"id": plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="الخطة غير موجودة")

    PROOF_DIR = "/app/uploads/payment_proofs"
    os.makedirs(PROOF_DIR, exist_ok=True)
    ext = os.path.splitext(proof.filename or "")[1].lower() or ".jpg"
    data = await proof.read()
    fname = f"inst_{plan_id}_{installment_number}_{uuid.uuid4().hex[:8]}{ext}"
    fpath = os.path.join(PROOF_DIR, fname)
    with open(fpath, "wb") as f:
        f.write(data)

    image_url = f"/api/files/payment_proofs/{fname}"

    # Update installment with proof
    plan_data = await db.installment_plans.find_one({"id": plan_id}, {"_id": 0})
    installments = plan_data.get("installments", [])
    for i, inst in enumerate(installments):
        if inst["number"] == installment_number:
            installments[i]["proof_url"] = image_url
            installments[i]["proof_notes"] = notes
            installments[i]["proof_uploaded_at"] = datetime.now(timezone.utc).isoformat()
            break

    await db.installment_plans.update_one(
        {"id": plan_id},
        {"$set": {"installments": installments}}
    )

    return {"success": True, "image_url": image_url}


@router.get("/financial/compound-summary")
async def get_compound_financial_summary(
    compound_id: str = None,
    current_user: dict = Depends(get_current_user),
):
    """إجماليات مالية شاملة للكمبوند — للداشبورد."""
    db = get_db()
    cid = compound_id or current_user.get("compound_id")
    if not cid:
        raise HTTPException(status_code=400, detail="compound_id مطلوب")

    # Installment plans totals
    plans = await db.installment_plans.find({"compound_id": cid}, {"_id": 0}).to_list(1000)
    total_installments = sum(p.get("total_amount", 0) for p in plans)
    total_collected = sum(
        i.get("paid_amount", 0)
        for p in plans for i in p.get("installments", [])
        if i.get("status") == "paid"
    )
    total_overdue_amount = sum(
        i.get("amount", 0)
        for p in plans for i in p.get("installments", [])
        if i.get("status") != "paid" and
        i.get("due_date", "") < datetime.now(timezone.utc).isoformat()
    )
    total_deposits = sum(p.get("deposit_amount", 0) for p in plans)
    overdue_residents = len(set(
        p["resident_id"] for p in plans
        for i in p.get("installments", [])
        if i.get("status") != "paid" and
        i.get("due_date", "") < datetime.now(timezone.utc).isoformat()
    ))

    # Unit charges
    total_charges = await db.unit_charges.aggregate([
        {"$match": {"compound_id": cid}},
        {"$group": {"_id": "$status", "total": {"$sum": "$amount"}}}
    ]).to_list(10)
    charges_by_status = {r["_id"]: r["total"] for r in total_charges}

    # Financial expenses/revenue
    from datetime import datetime, timezone
    month_start = datetime.now(timezone.utc).replace(day=1).isoformat()
    month_expenses = await db.expenses.aggregate([
        {"$match": {"compound_id": cid, "created_at": {"$gte": month_start}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    month_revenue = await db.revenue.aggregate([
        {"$match": {"compound_id": cid, "created_at": {"$gte": month_start}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)

    return {
        "compound_id": cid,
        "installments": {
            "total_amount": total_installments,
            "collected": total_collected,
            "remaining": total_installments - total_collected,
            "overdue_amount": total_overdue_amount,
            "overdue_residents": overdue_residents,
            "total_deposits": total_deposits,
            "plans_count": len(plans),
        },
        "charges": {
            "pending": charges_by_status.get("pending", 0),
            "paid": charges_by_status.get("paid", 0),
            "overdue": charges_by_status.get("overdue", 0),
        },
        "this_month": {
            "expenses": month_expenses[0]["total"] if month_expenses else 0,
            "revenue": month_revenue[0]["total"] if month_revenue else 0,
        },
        "collection_rate": round(
            (total_collected / total_installments * 100) if total_installments > 0 else 0, 1
        ),
    }

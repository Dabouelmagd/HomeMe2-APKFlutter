"""
Test Excel Export and Daily Reports Features
- Excel export endpoint with 5 sheets (Balance Sheet, Expenses, Unit Charges, Obligations, Revenue)
- Trigger daily reports endpoint
- Balance sheet monthly_breakdown for comparison tab
"""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestExcelExportAndDailyReports:
    """Test Excel export and daily report features"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup - login and get token"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        # Login as admin
        login_response = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "username": "admin",
            "password": "admin123"
        })
        
        if login_response.status_code == 200:
            data = login_response.json()
            token = data.get("access_token") or data.get("token")
            self.session.headers.update({"Authorization": f"Bearer {token}"})
            self.token = token
        else:
            pytest.skip(f"Login failed: {login_response.status_code}")
    
    # ==================== EXCEL EXPORT TESTS ====================
    
    def test_excel_export_returns_valid_file(self):
        """Test GET /api/financial/export-excel returns valid Excel file"""
        response = self.session.get(f"{BASE_URL}/api/financial/export-excel?year=2025")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        # Check content type
        content_type = response.headers.get('Content-Type', '')
        assert 'spreadsheet' in content_type or 'application/vnd.openxmlformats' in content_type, \
            f"Expected Excel content type, got: {content_type}"
        
        # Check content disposition header
        content_disp = response.headers.get('Content-Disposition', '')
        assert 'attachment' in content_disp, f"Expected attachment header, got: {content_disp}"
        assert '.xlsx' in content_disp, f"Expected .xlsx filename, got: {content_disp}"
        
        # Check file size is reasonable (not empty)
        assert len(response.content) > 1000, f"Excel file too small: {len(response.content)} bytes"
        
        print(f"✓ Excel export successful - file size: {len(response.content)} bytes")
    
    def test_excel_export_with_month_filter(self):
        """Test Excel export with month filter"""
        response = self.session.get(f"{BASE_URL}/api/financial/export-excel?year=2025&month=1")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert len(response.content) > 500, "Excel file should have content"
        
        print(f"✓ Excel export with month filter successful")
    
    def test_excel_export_has_5_sheets(self):
        """Test Excel file contains 5 sheets: Balance Sheet, Expenses, Unit Charges, Obligations, Revenue"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        response = self.session.get(f"{BASE_URL}/api/financial/export-excel?year=2025")
        assert response.status_code == 200
        
        # Load workbook from response content
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        # Check for expected sheets
        expected_sheets = ["Balance Sheet", "Expenses", "Unit Charges", "Obligations", "Revenue"]
        for expected in expected_sheets:
            assert expected in sheet_names, f"Missing sheet: {expected}. Found: {sheet_names}"
        
        print(f"✓ Excel has all 5 sheets: {sheet_names}")
        
        # Verify each sheet has headers
        for sheet_name in expected_sheets:
            ws = wb[sheet_name]
            # Check first row has content (headers)
            first_row = [cell.value for cell in ws[1]]
            assert any(first_row), f"Sheet {sheet_name} has no headers"
            print(f"  - {sheet_name}: headers = {first_row[:3]}...")
    
    def test_excel_export_requires_auth(self):
        """Test Excel export requires authentication"""
        # Create new session without auth
        no_auth_session = requests.Session()
        response = no_auth_session.get(f"{BASE_URL}/api/financial/export-excel?year=2025")
        
        assert response.status_code in [401, 403], f"Expected 401/403 without auth, got {response.status_code}"
        print("✓ Excel export correctly requires authentication")
    
    # ==================== DAILY REPORTS TESTS ====================
    
    def test_trigger_daily_reports_endpoint(self):
        """Test POST /api/email/trigger-daily-reports sends reports"""
        response = self.session.post(f"{BASE_URL}/api/email/trigger-daily-reports")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        data = response.json()
        assert "message" in data, f"Response should have message field: {data}"
        assert "emails_sent" in data, f"Response should have emails_sent field: {data}"
        
        # emails_sent should be a number
        assert isinstance(data["emails_sent"], int), f"emails_sent should be int: {data['emails_sent']}"
        
        print(f"✓ Trigger daily reports successful - {data['emails_sent']} emails sent")
        print(f"  Message: {data['message']}")
    
    def test_trigger_daily_reports_requires_admin(self):
        """Test trigger daily reports requires admin role"""
        # Create new session without auth
        no_auth_session = requests.Session()
        response = no_auth_session.post(f"{BASE_URL}/api/email/trigger-daily-reports")
        
        assert response.status_code in [401, 403], f"Expected 401/403 without auth, got {response.status_code}"
        print("✓ Trigger daily reports correctly requires admin authentication")
    
    # ==================== BALANCE SHEET MONTHLY BREAKDOWN TESTS ====================
    
    def test_balance_sheet_returns_monthly_breakdown(self):
        """Test GET /api/financial/balance-sheet returns monthly_breakdown for comparison tab"""
        response = self.session.get(f"{BASE_URL}/api/financial/balance-sheet?year=2025")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        data = response.json()
        
        # Check required fields for comparison tab
        assert "monthly_breakdown" in data, f"Response should have monthly_breakdown: {data.keys()}"
        
        # monthly_breakdown should be a dict
        monthly = data["monthly_breakdown"]
        assert isinstance(monthly, dict), f"monthly_breakdown should be dict: {type(monthly)}"
        
        print(f"✓ Balance sheet has monthly_breakdown with {len(monthly)} months")
        
        # If there's data, check structure
        if monthly:
            sample_month = list(monthly.keys())[0]
            sample_data = monthly[sample_month]
            print(f"  Sample month {sample_month}: {sample_data}")
            
            # Each month should have expenses and revenue
            assert "expenses" in sample_data or "revenue" in sample_data, \
                f"Month data should have expenses/revenue: {sample_data}"
    
    def test_balance_sheet_has_collection_rate(self):
        """Test balance sheet includes collection rate for comparison"""
        response = self.session.get(f"{BASE_URL}/api/financial/balance-sheet?year=2025")
        
        assert response.status_code == 200
        data = response.json()
        
        # Check obligations section has collection_rate
        if "obligations" in data:
            obl = data["obligations"]
            assert "collection_rate" in obl, f"Obligations should have collection_rate: {obl.keys()}"
            print(f"✓ Collection rate: {obl['collection_rate']}%")
        else:
            print("⚠ No obligations data in balance sheet")
    
    def test_balance_sheet_has_expenses_by_category(self):
        """Test balance sheet includes expenses_by_category for pie chart"""
        response = self.session.get(f"{BASE_URL}/api/financial/balance-sheet?year=2025")
        
        assert response.status_code == 200
        data = response.json()
        
        assert "expenses_by_category" in data, f"Should have expenses_by_category: {data.keys()}"
        print(f"✓ Expenses by category: {data['expenses_by_category']}")
    
    def test_balance_sheet_has_revenue_by_source(self):
        """Test balance sheet includes revenue_by_source for pie chart"""
        response = self.session.get(f"{BASE_URL}/api/financial/balance-sheet?year=2025")
        
        assert response.status_code == 200
        data = response.json()
        
        assert "revenue_by_source" in data, f"Should have revenue_by_source: {data.keys()}"
        print(f"✓ Revenue by source: {data['revenue_by_source']}")


class TestComparisonTabData:
    """Test data requirements for monthly comparison tab"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup - login and get token"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        login_response = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "username": "admin",
            "password": "admin123"
        })
        
        if login_response.status_code == 200:
            data = login_response.json()
            token = data.get("access_token") or data.get("token")
            self.session.headers.update({"Authorization": f"Bearer {token}"})
        else:
            pytest.skip(f"Login failed: {login_response.status_code}")
    
    def test_monthly_data_structure_for_charts(self):
        """Test monthly breakdown has correct structure for Recharts"""
        response = self.session.get(f"{BASE_URL}/api/financial/balance-sheet?year=2025")
        
        assert response.status_code == 200
        data = response.json()
        
        monthly = data.get("monthly_breakdown", {})
        
        # Verify structure is suitable for charts
        for month_key, month_data in monthly.items():
            # Month key should be in format YYYY-MM
            assert "-" in month_key, f"Month key should be YYYY-MM format: {month_key}"
            
            # Data should have numeric values
            if "expenses" in month_data:
                assert isinstance(month_data["expenses"], (int, float)), \
                    f"Expenses should be numeric: {month_data['expenses']}"
            if "revenue" in month_data:
                assert isinstance(month_data["revenue"], (int, float)), \
                    f"Revenue should be numeric: {month_data['revenue']}"
        
        print(f"✓ Monthly data structure valid for charts")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
